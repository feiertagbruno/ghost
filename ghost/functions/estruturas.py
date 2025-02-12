from dotenv import load_dotenv
from os import environ, path
from sqlalchemy import create_engine, text
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from ghost.queries import *
from django.conf import settings
from django.contrib import messages
from ghost.models import Processamento
from ghost.utils.funcs import get_last_day_of_month


def tratamento_data_referencia(data_referencia):
	if not data_referencia:
		data_referencia = datetime.today().date()
	elif type(data_referencia) == str:
		if data_referencia.count("-") == 2:
			data_referencia = datetime.strptime(data_referencia, "%Y-%m-%d").date()
		elif data_referencia.count("/") == 2:
			data_referencia = datetime.strptime(data_referencia, "%d/%m/%Y").date()
		else:
			data_referencia = datetime.strptime(data_referencia, "%Y%m%d").date()
	return data_referencia




def get_engine():
	load_dotenv()

	SERVER = environ.get("SERVER")
	DB = environ.get("DB")
	USER = environ.get("USER")
	PWD = environ.get("PWD")
	DRIVER = "ODBC Driver 18 for SQL Server"

	connection_string = f"mssql+pyodbc://{USER}:{PWD}@{SERVER}/{DB}?driver={DRIVER}&TrustServerCertificate=yes"

	engine = create_engine(connection_string)
	
	return engine




def get_estrutura_produto(codigo, data_referencia, engine):

	query = get_query_estrutura_produto()
	resultado = pd.read_sql(text(query), engine, params={
		"data_referencia": data_referencia,
		"codigo": codigo,
	})

	return resultado




def forma_string_codigos(codigos):
	str_codigos = ""
	for cod in codigos:
		str_codigos += f"{cod},"
	return str_codigos.rstrip(",")





def acrescenta_alternativos(estrutura: pd.DataFrame, engine, abre_todos_os_PIs = True):
	if not abre_todos_os_PIs:
		codigos_filhos = estrutura[(estrutura["tipo_insumo"] != "PI") | (
			(estrutura["tipo_insumo"] == "PI") & (estrutura["fantasma"] != "S")
		)][["insumo"]].reset_index(drop=True)
	else:
		codigos_filhos = estrutura[estrutura["tipo_insumo"] != "PI"][["insumo"]].reset_index(drop=True)
	todos_os_codigos = codigos_filhos.copy(deep=True).rename(columns={"insumo":"todos_os_codigos"})

	str_codigos = forma_string_codigos(codigos_filhos["insumo"])

	query = get_query_alternativos()

	resultado = pd.read_sql(text(query), engine, params={
		"codigos": str_codigos,
	})

	todos_os_codigos = pd.concat([todos_os_codigos,(
		resultado.drop_duplicates(subset="alternativos")[["alternativos"]].rename(columns={"alternativos":"todos_os_codigos"})
	)],ignore_index=True)
	todos_os_codigos.drop_duplicates(subset="todos_os_codigos",inplace=True)

	resultado = resultado.groupby("prodori")["alternativos"].agg(lambda x: ";".join(x)).reset_index()
	estrutura = estrutura.merge(resultado[["prodori","alternativos"]],how="left",left_on="insumo",right_on="prodori")\
		.drop(columns="prodori").fillna(value="")

	return estrutura, todos_os_codigos




def busca_custos_ultima_compra(str_codigos, data_referencia, engine):

	query = get_query_ultima_compra_produtos()

	return pd.read_sql(text(query), engine, params={
		"codigos":str_codigos,
		"data_referencia": data_referencia,
	})

	#começar pelo preço da última compra




def explode_extrutura(codigo, data_referencia = None, engine = None, abre_todos_os_PIs = True):

	data_referencia = tratamento_data_referencia(data_referencia)
	
	if not engine:
		engine = get_engine()

	#primeira consulta no banco
	resultado = get_estrutura_produto(codigo, data_referencia, engine)
	estrutura = resultado.copy(deep=True)

	if not abre_todos_os_PIs:
		filtro_PI = resultado[(resultado["tipo_insumo"] == "PI") & (resultado["fantasma"] == "S")]
	else:
		filtro_PI = resultado[resultado["tipo_insumo"] == "PI"]
	tem_PI = True if not filtro_PI.empty else False

	#enquanto tiver PI continua executando a consulta no banco
	while tem_PI:
		
		for codigo_PI in filtro_PI["insumo"]:
			if codigo_PI == 'MPBPJ0000000586':
				a = ''
			resultado = get_estrutura_produto(codigo_PI, data_referencia, engine)
			estrutura.loc[estrutura["insumo"] == codigo_PI, "verificado"] = True
			estrutura = pd.concat([estrutura,resultado],ignore_index=True)

		if not abre_todos_os_PIs:
			filtro_PI = estrutura[(estrutura["tipo_insumo"] == "PI") & (estrutura["fantasma"] == "S") & (estrutura["verificado"] != True)]
		else:
			filtro_PI = estrutura[(estrutura["tipo_insumo"] == "PI") & (estrutura["verificado"] != True)]
		tem_PI = True if not filtro_PI.empty else False

	resultado = None
	filtro_PI = None
	tem_PI = None

	estrutura["quant_utilizada"] = estrutura.groupby("insumo")["quant_utilizada"].transform("sum")
	estrutura.drop_duplicates(subset="insumo", inplace=True)
	
	estrutura, todos_os_codigos = acrescenta_alternativos(estrutura, engine, abre_todos_os_PIs)

	return estrutura, todos_os_codigos

	# FIM explode_extrutura




def busca_custo_e_comentario_alternativos(
		row, consulta_custos, coluna_custo, coluna_custo_alt, coluna_comentario, coluna_comentario_alt):
	alternativos = row["alternativos"].split(";")
	
	if not alternativos[0]: 
		return pd.Series({
		coluna_custo_alt: "",
		coluna_comentario_alt: ""
	})

	custos_alternativos = [
		f'{round(consulta_custos.loc[consulta_custos["insumo"] == alt, coluna_custo].values[0] * row["quant_utilizada"],5):.5f}'.replace(".",",")
		for alt in alternativos if alt in consulta_custos["insumo"].values
	]
	comentarios_alternativos = [
		consulta_custos.loc[consulta_custos["insumo"] == alt, coluna_comentario].values[0]
		for alt in alternativos if alt in consulta_custos["insumo"].values
	]

	return pd.Series({
		coluna_custo_alt: ";".join(custos_alternativos) if custos_alternativos != [None] else "",
		coluna_comentario_alt: "\n\n".join(comentarios_alternativos) if comentarios_alternativos != [None] else ""
	})




def traz_custos_por_produto(estrutura, consulta_custos, tupla_nomes_colunas):
	coluna_custo, coluna_comentario = tupla_nomes_colunas

	estrutura = estrutura.merge(consulta_custos[["insumo",coluna_custo,coluna_comentario]], on="insumo", how="left")\
		.assign(**{coluna_custo: lambda x: x["quant_utilizada"] * x[coluna_custo]}).fillna(0)

	if "alternativos" in estrutura.columns:

		coluna_custo_alt = f"{coluna_custo}_alt"
		coluna_comentario_alt = f"{coluna_comentario}_alt"
		estrutura[coluna_custo_alt] = ""
		estrutura[coluna_comentario_alt] = ""

		estrutura[[coluna_custo_alt,coluna_comentario_alt]] = estrutura.apply(
			busca_custo_e_comentario_alternativos, 
			axis=1, 
			args=(consulta_custos, coluna_custo, coluna_custo_alt, coluna_comentario, coluna_comentario_alt)
		)
	
	return estrutura




def calcula_custo_total(codigo, descricao, data_referencia_var, estrutura:pd.DataFrame, 
						nomes_colunas, custos_totais_produto = pd.DataFrame()):
	coluna_custos_estrutura, coluna_custos_alternativos, nova_coluna_custo_total, nova_coluna_comentario = nomes_colunas
	
	if custos_totais_produto.empty:
		custos_totais_produto = pd.DataFrame([{
			"codigo_original": codigo,
			"descricao_cod_original": descricao,
			"data_referencia": pd.to_datetime(data_referencia_var, format="%d/%m/%Y"),
			nova_coluna_custo_total: None,
			nova_coluna_comentario: None,
		}])
	elif custos_totais_produto[custos_totais_produto["codigo_original"] == codigo].empty:
		custos_totais_produto.loc[len(custos_totais_produto)] = [
			codigo, descricao, data_referencia_var
		]

	comentario = f"{codigo}\nData Referência: {datetime.strftime(data_referencia_var,'%d/%m/%Y')}\n\n"
	#calculo dos originais (os únicos no caso de OPs)
	estrutura_codigos_originais = estrutura[
		(estrutura[coluna_custos_estrutura].notnull()) & (estrutura[coluna_custos_estrutura] != 0)
	][["insumo","descricao_insumo",coluna_custos_estrutura]]

	if not estrutura_codigos_originais.empty:
		total = estrutura_codigos_originais[coluna_custos_estrutura].sum()
		comentario += estrutura_codigos_originais[["insumo","descricao_insumo",coluna_custos_estrutura]].agg(
			lambda x: f"Ori - {x['insumo']} - {x['descricao_insumo']} - R$ {str(round(x[coluna_custos_estrutura],5)).replace('.',',')}"
			,axis=1
		).str.cat(sep="\n")
	else:
		total = 0
		comentario += ""

	produtos_nao_encontrados = estrutura[
		(estrutura[coluna_custos_estrutura].isnull()) | (estrutura[coluna_custos_estrutura] == 0)
	][["insumo","descricao_insumo"]]


	#quando tem a coluna alternativos
	if "alternativos" in estrutura.columns:
		estrutura_alternativos = estrutura[(
			((estrutura[coluna_custos_estrutura].isnull()) | (estrutura[coluna_custos_estrutura] == 0)) &
			((estrutura[coluna_custos_alternativos].notnull()) & (estrutura[coluna_custos_alternativos] != ""))
		)][["alternativos", coluna_custos_alternativos]]
		produtos_nao_encontrados = estrutura[(
			((estrutura[coluna_custos_estrutura].isnull()) | (estrutura[coluna_custos_estrutura] == 0)) &
			((estrutura[coluna_custos_alternativos].isnull()) | (estrutura[coluna_custos_alternativos] == "")) & 
			((estrutura["alternativos"].notnull()) & estrutura["alternativos"] != "")
		)][["insumo","descricao_insumo"]]

		if not estrutura_alternativos.empty:
			total += estrutura_alternativos[coluna_custos_alternativos].str.split(";").str[0]\
				.str.replace(",",".").astype(float).sum()
			# código inserido para quando o original for zero ou vazio e o alternativo tiver valor,
			# o valor do alternativo substituir o 0 que está no original
			estrutura = estrutura.merge(
				estrutura_alternativos.rename(columns={coluna_custos_alternativos:"custo_alt"}), 
				on="alternativos", 
				how="left"
			)
			estrutura[coluna_custos_estrutura] = \
				estrutura["custo_alt"].str.split(";").str[0].str.replace(",",".").astype(float)\
					.combine_first(estrutura[coluna_custos_estrutura])
			estrutura = estrutura.drop(columns=["custo_alt"])
			#
			
			comentario += estrutura_alternativos[["alternativos",coluna_custos_alternativos]].agg(
				lambda x: f"Alt-{x['alternativos'].split(';')[0]}: R$ {x[coluna_custos_alternativos].split(';')[0]}"
				,axis=1
			).str.cat(sep="\n")

	#produtos não encontrados
	comentario += "\n\nProdutos não encontrados:\n"
	produtos_nao_encontrados[["insumo", "descricao_insumo"]] = produtos_nao_encontrados[["insumo", "descricao_insumo"]].astype(str)
	comentario += (produtos_nao_encontrados["insumo"] + " - " + produtos_nao_encontrados["descricao_insumo"]).str.cat(sep="\n")
	
	custos_totais_produto.loc[custos_totais_produto["codigo_original"] == codigo, nova_coluna_custo_total] = round(total,5)
	custos_totais_produto.loc[custos_totais_produto["codigo_original"] == codigo, nova_coluna_comentario] = comentario

	return estrutura, custos_totais_produto




def busca_custos_ultimo_fechamento(str_codigos, data_referencia, engine):
	query = get_query_ultimo_fechamento_produtos()
	
	return pd.read_sql(text(query), engine, params={
		"codigos": str_codigos,
		"data_referencia": data_referencia,
	})




def busca_custos_medios(str_codigos, data_referencia, engine):
	query = get_query_custos_medios_produtos()
	
	return pd.read_sql(text(query), engine, params={"codigos":str_codigos})




def get_descricao_produto(codigo, engine = None):
	return pd.read_sql(
			text(get_query_busca_descricao_produto()),
			engine, params={"codigo":codigo}
		)["descricao"].values[0]




def busca_compra_mais_antiga_por_data_ref(str_codigos, data_referencia, engine):
	if not engine:
		engine = get_engine()
	
	query = get_query_compra_mais_antiga()

	return pd.read_sql(text(query), engine, params={
		"produtos": str_codigos,
		"data_referencia": data_referencia,
	})



def busca_menor_fechamento_por_data_ref(str_codigos, data_referencia, engine):
	if not engine:
		engine = get_engine()
	
	query = get_query_menor_fechamento()

	return pd.read_sql(text(query),engine,params={
		"produtos": str_codigos,
		"data_fechamento": data_referencia
	})




def estrutura_simples(codigo, data_referencia, engine = None, abre_todos_os_PIs = True, data_std = None):

	if not engine:
		engine = get_engine()
	data_referencia = tratamento_data_referencia(data_referencia)
	if data_std:
		if isinstance(data_std, str):
			data_std = get_last_day_of_month(data_std)

	estrutura, todos_os_codigos = explode_extrutura(
		codigo, 
		data_std if data_std else data_referencia, 
		engine, 
		abre_todos_os_PIs
	)

	if not abre_todos_os_PIs:
		estrutura = estrutura[(estrutura["tipo_insumo"] != "PI") | \
						((estrutura["tipo_insumo"] == "PI") & (estrutura["fantasma"] != "S"))]
	else:
		estrutura = estrutura[estrutura["tipo_insumo"] != "PI"]

	str_codigos = forma_string_codigos(todos_os_codigos["todos_os_codigos"])

	try:
		descricao = estrutura.loc[estrutura["codigo_pai"] == codigo, "descricao_pai"].values[0]
	except:
		descricao = get_descricao_produto(codigo, engine)

	# ÚLTIMA COMPRA
	custos_ultima_compra = busca_custos_ultima_compra(
		str_codigos, 
		data_std if data_std else data_referencia, 
		engine
	)
	custos_ultima_compra = custos_ultima_compra[
		(custos_ultima_compra["ult_compra_custo_utilizado"] != 0) & 
		(custos_ultima_compra["ult_compra_custo_utilizado"].notnull())
	]
	# TRAZER CUSTOS OLHANDO PARA FRENTE NO CASO DO BOMXOPSTD
	if data_std:
		codigos_nao_encontrados = todos_os_codigos[
			~todos_os_codigos["todos_os_codigos"].isin(custos_ultima_compra["insumo"])
		]
		if not codigos_nao_encontrados.empty:
			codigos_nao_encontrados_str = forma_string_codigos(codigos_nao_encontrados["todos_os_codigos"])
			codigos_nao_encontrados = busca_compra_mais_antiga_por_data_ref(
				codigos_nao_encontrados_str, 
				data_std if data_std else data_referencia, 
				engine
			)
			custos_ultima_compra = pd.concat([custos_ultima_compra,codigos_nao_encontrados],axis=0, ignore_index=True)
			codigos_nao_encontrados = None

	###
	estrutura = traz_custos_por_produto(estrutura, custos_ultima_compra, 
									 ("ult_compra_custo_utilizado","comentario_ultima_compra"))
	estrutura, custos_totais_produto = calcula_custo_total(
		codigo, descricao, data_std if data_std else data_referencia, estrutura, 
		[
			"ult_compra_custo_utilizado",
			"ult_compra_custo_utilizado_alt",
			"custo_total_ultima_compra",
			"comentario_ultima_compra"
		])
	
	# FECHAMENTO
	custos_ultimo_fechamento = busca_custos_ultimo_fechamento(
		str_codigos, 
		data_std if data_std else data_referencia, 
		engine
	)

	# TRAZER CUSTOS OLHANDO PARA FRENTE NO CASO DO BOMXOPSTD
	if data_std:
		codigos_nao_encontrados = todos_os_codigos[
			~todos_os_codigos["todos_os_codigos"].isin(custos_ultimo_fechamento["insumo"])
		]
		if not codigos_nao_encontrados.empty:
			codigos_nao_encontrados_str = forma_string_codigos(codigos_nao_encontrados["todos_os_codigos"])
			codigos_nao_encontrados = busca_menor_fechamento_por_data_ref(
				codigos_nao_encontrados_str, 
				data_std if data_std else data_referencia, 
				engine
			)
			custos_ultimo_fechamento = pd.concat(
				[custos_ultimo_fechamento,codigos_nao_encontrados],axis=0, ignore_index=True
			)

	###
	
	estrutura = traz_custos_por_produto(estrutura, custos_ultimo_fechamento, 
									 ("fechamento_custo_utilizado","comentario_fechamento"))
	estrutura, custos_totais_produto = calcula_custo_total(
		codigo, descricao, data_std if data_std else data_referencia, estrutura,
		[
			"fechamento_custo_utilizado",
			"fechamento_custo_utilizado_alt",
			"custo_total_ultimo_fechamento",
			"comentario_ultimo_fechamento"
		],
		custos_totais_produto)
	
	# MÉDIO ATUAL
	custos_medios = busca_custos_medios(str_codigos, data_referencia, engine)
	estrutura = traz_custos_por_produto(estrutura, custos_medios, 
									 ("medio_atual_custo_utilizado", "comentario_custo_medio"))
	estrutura, custos_totais_produto = calcula_custo_total(codigo, descricao, data_referencia, estrutura, 
		[
			"medio_atual_custo_utilizado",
			"medio_atual_custo_utilizado_alt",
			"total_pelo_custo_medio", 
			"comentario_custo_medio"
		],
		custos_totais_produto)
	
	estrutura["codigo_original"] = str(codigo)
	estrutura["descricao_cod_original"] = str(descricao)

	return estrutura, custos_totais_produto
	
def gerar_relatorio_excel_estruturas_simples(estrutura, custos_totais_produtos, data_referencia):

	data_referencia = tratamento_data_referencia(data_referencia)

	max_lengths = {}
	for col in estrutura.columns:
		max_length = max(estrutura[col].apply(lambda x: len(str(x))).max(), len(str(col))) 
		max_lengths[col] = max_length

	nomes_das_colunas = {
		'codigo_original':"Cód Original",
		'descricao_cod_original':"Desc Orig",
		'codigo_pai':"Código Pai", 
		'descricao_pai':"Desc Pai", 
		'tipo_pai':"Tipo Pai", 
		'insumo':"Insumo", 
		'descricao_insumo':"Descrição Insumo",
		'quant_utilizada':"Quant Utilizada", 
		'tipo_insumo':"Tipo Insumo", 
		'origem':"Origem", 
		'alternativos':"Alternativos", 
		'ult_compra_custo_utilizado':"Últ Compra",
		'ult_compra_custo_utilizado_alt':"Últ Compra Alt",
		'fechamento_custo_utilizado':"Últ Fechamento",
		'fechamento_custo_utilizado_alt':"Últ Fecham Alt",
		'medio_atual_custo_utilizado':"Médio Atual",
		'medio_atual_custo_utilizado_alt':"Médio Atual Alt",
	}

	wb = Workbook()
	ws = wb.active
	ws.title = "Detalhamento"

	ws.append([
		nomes_das_colunas["codigo_original"],
		nomes_das_colunas["descricao_cod_original"],
		nomes_das_colunas["codigo_pai"],
		nomes_das_colunas["descricao_pai"],
		nomes_das_colunas["tipo_pai"],
		nomes_das_colunas["insumo"],
		nomes_das_colunas["descricao_insumo"],
		nomes_das_colunas["quant_utilizada"],
		nomes_das_colunas["tipo_insumo"],
		nomes_das_colunas["origem"],
		nomes_das_colunas["alternativos"],
		nomes_das_colunas["ult_compra_custo_utilizado"],
		nomes_das_colunas["ult_compra_custo_utilizado_alt"],
		nomes_das_colunas["fechamento_custo_utilizado"],
		nomes_das_colunas["fechamento_custo_utilizado_alt"],
		nomes_das_colunas["medio_atual_custo_utilizado"],
		nomes_das_colunas["medio_atual_custo_utilizado_alt"],
	])

	for i, row in estrutura.iterrows():
		l = i+2
		ws.cell(l, 1, row["codigo_original"])
		ws.cell(l, 2, row["descricao_cod_original"])
		ws.cell(l, 3, row["codigo_pai"])
		ws.cell(l, 4, row["descricao_pai"])
		ws.cell(l, 5, row["tipo_pai"])
		ws.cell(l, 6, row["insumo"])
		ws.cell(l, 7, row["descricao_insumo"])
		ws.cell(l, 8, row["quant_utilizada"])
		ws.cell(l, 9, row["tipo_insumo"])
		ws.cell(l,10, row["origem"])
		ws.cell(l,11, row["alternativos"])

		comm = row["comentario_ultima_compra"]
		ws.cell(l,12, row["ult_compra_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_ultima_compra_alt"]
		ws.cell(l,13, row["ult_compra_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_fechamento"]
		ws.cell(l,14, row["fechamento_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_fechamento_alt"]
		ws.cell(l,15, row["fechamento_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio"]
		ws.cell(l,16, row["medio_atual_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_custo_medio_alt"]
		ws.cell(l,17, row["medio_atual_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

	colunas_para_ocultar = {
		"Desc Orig","Desc Pai","Descrição Insumo"
	}

	for col in range(1, ws.max_column + 1):
		cabecalho = ws.cell(1,col).value
		letra_coluna = get_column_letter(col)

		if cabecalho in colunas_para_ocultar:
			ws.column_dimensions[letra_coluna].outlineLevel = 1
			ws.column_dimensions[letra_coluna].hidden = True
		
		for chave, valor in nomes_das_colunas.items():
			if valor == cabecalho:
				adjusted_width = max_lengths[chave] + 4
				ws.column_dimensions[letra_coluna].width = adjusted_width
				break

	tab = Table(
		displayName="tabela_estruturas",
		ref=f"A1:Q{len(estrutura) + 1}"
	)
	style = TableStyleInfo(
		name="TableStyleMedium2",
		showFirstColumn=False,
		showLastColumn=False,
		showRowStripes=True,
		showColumnStripes=False,
	)
	tab.tableStyleInfo = style
	ws.add_table(tab)

	ws2 = wb.create_sheet("Consolidado",0)
	ws2.append(["Data de Referência", "Código","Descrição","Últimas Entradas","Último Fechamento","Custo Médio"])

	for i, row in custos_totais_produtos.iterrows():
		l = i+2
		ws2.cell(l, 1, data_referencia).number_format = "DD/MM/YYYY"
		ws2.cell(l, 2, row["codigo_original"])
		ws2.cell(l, 3, row["descricao_cod_original"])

		comm = row["comentario_ultima_compra"]
		ws2.cell(l, 4, row["custo_total_ultima_compra"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_ultimo_fechamento"]
		ws2.cell(l, 5, row["custo_total_ultimo_fechamento"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio"]
		ws2.cell(l, 6, row["total_pelo_custo_medio"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

	tab = Table(
		displayName="tabela_consolidada",
		ref=f"A1:F{len(custos_totais_produtos) + 1}"
	)
	tab.tableStyleInfo = style
	ws2.add_table(tab)

	caminho_arquivo = path.join(settings.MEDIA_ROOT,"Estruturas.xlsx")

	wb.save(caminho_arquivo)

	return caminho_arquivo




def gerar_multiestruturas(
		request, produtos, data_referencia, engine = None
	):

	if not engine:
		engine = get_engine()

	compilado_estruturas = pd.DataFrame()
	compilado_custos_totais = pd.DataFrame()

	data_referencia = tratamento_data_referencia(data_referencia)

	################ processamento
	processamento = Processamento.objects.create(
		codigo_identificador = request.POST.get("codigo-identificador","a"),
		caller = "multiestruturas",
		porcentagem = "",
		mensagem1 = "Processando",
		mensagem2 = ""
	)
	################

	for index, produto in enumerate(produtos):
		produto = produto.strip().upper()

		################ processamento
		processamento.porcentagem = f'{int((index+1)/len(produtos)*98)}%'
		processamento.mensagem2 = f'Extraindo estrutura de {produto}'
		processamento.save()
		################

		if len(produto) == 7 or len(produto) == 15:
			estrutura, custos_totais_produto = estrutura_simples(produto, data_referencia, engine)
			compilado_estruturas = pd.concat([compilado_estruturas, estrutura])
			compilado_custos_totais = pd.concat([compilado_custos_totais, custos_totais_produto])
		else:
			messages.info(request, f"Produto {produto} contém um erro de digitação.")
	
	compilado_custos_totais = compilado_custos_totais.reset_index(drop=True)
	compilado_estruturas = compilado_estruturas.reset_index(drop=True)

	return [request, compilado_estruturas, compilado_custos_totais]