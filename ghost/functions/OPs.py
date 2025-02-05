from django.conf import settings
from ghost.functions.estruturas import (
	get_engine, tratamento_data_referencia, forma_string_codigos,
	busca_custos_ultima_compra, busca_custos_ultimo_fechamento, busca_custos_medios,
	traz_custos_por_produto, calcula_custo_total, get_descricao_produto
)
from ghost.queries import (
	get_query_detalhamento_op,get_query_numeros_op_por_periodo, get_query_busca_op_pelo_produto
)
import pandas as pd
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from os import path


def get_info_op(numero_op, engine = None):

	if not engine:
		engine = get_engine()

	query = get_query_detalhamento_op()

	resultado = pd.read_sql(text(query), engine, params={"numero_op": numero_op} )

	if resultado.empty: return [None, None, None, None]

	codigo = resultado["codigo_original"].values[0]
	data_referencia = resultado["data_referencia"].values[0]
	data_referencia = tratamento_data_referencia(data_referencia)
	descricao = get_descricao_produto(codigo, engine)

	str_codigos = forma_string_codigos(resultado["insumo"])

	custos_ultima_compra = busca_custos_ultima_compra(str_codigos, data_referencia, engine)
	resultado = traz_custos_por_produto(resultado, custos_ultima_compra, ("ult_compra_custo_utilizado","comentario_ultima_compra"))
	custos_totais_op = calcula_custo_total(codigo, descricao, data_referencia, resultado,
		["ult_compra_custo_utilizado", "", "custo_total_ultima_compra_op","comentario_ultima_compra_op"]
	)

	custos_ultimo_fechamento = busca_custos_ultimo_fechamento(str_codigos, data_referencia, engine)
	resultado = traz_custos_por_produto(resultado, custos_ultimo_fechamento, ("fechamento_custo_utilizado","comentario_fechamento"))
	custos_totais_op = calcula_custo_total(codigo, descricao, data_referencia, resultado,
		["fechamento_custo_utilizado", "", "custo_total_ult_fechamento_op", "comentario_fechamento_op"],
		custos_totais_op
	)

	custos_medios = busca_custos_medios(str_codigos, data_referencia, engine)
	resultado = traz_custos_por_produto(resultado, custos_medios, ("medio_atual_custo_utilizado", "comentario_custo_medio"))
	custos_totais_op = calcula_custo_total(codigo, descricao, data_referencia, resultado,
		["medio_atual_custo_utilizado", "", "total_pelo_custo_medio_op", "comentario_custo_medio_op"],
		custos_totais_op
	)

	custos_totais_op["op"] = numero_op

	return [codigo, data_referencia, resultado, custos_totais_op]




def linha_a_linha_estru_op(row, consulta_op):
	insumo = row["insumo"]
	consulta_filtrada = consulta_op[consulta_op["insumo"] == insumo]

	if consulta_filtrada.empty:
		alternativos = row["alternativos"]
		if alternativos and alternativos != "":
			alternativos = alternativos.split(";")
			for alternativo in alternativos:
				consulta_filtrada = consulta_op[consulta_op["insumo"] == alternativo]

				if not consulta_filtrada.empty:

					op = consulta_filtrada["op"].iloc[0]
					insumo_op = consulta_filtrada["insumo"].iloc[0]
					descricao_insumo_op = consulta_filtrada["descricao_insumo"].iloc[0]
					quant_utilizada_op = consulta_filtrada["quant_utilizada"].iloc[0]
					quant_produzida = consulta_filtrada["quant_produzida"].iloc[0]
					quant_total_utilizada = consulta_filtrada["quant_total_utilizada"].iloc[0]
					ult_compra_custo_utilizado_op = consulta_filtrada["ult_compra_custo_utilizado"].iloc[0]
					comentario_ultima_compra_op = consulta_filtrada["comentario_ultima_compra"].iloc[0]
					fechamento_custo_utilizado_op = consulta_filtrada["fechamento_custo_utilizado"].iloc[0]
					comentario_fechamento_op = consulta_filtrada["comentario_fechamento"].iloc[0]
					medio_atual_custo_utilizado_op = consulta_filtrada["medio_atual_custo_utilizado"].iloc[0]
					comentario_custo_medio_op = consulta_filtrada["comentario_custo_medio"].iloc[0]
					emissao_op = consulta_filtrada["data_referencia"].iloc[0]
					fechamento_op = consulta_filtrada["data_encerramento_op"].iloc[0]
					
					return pd.Series({
						"op": op, 
						"insumo_op": insumo_op,
						"descricao_insumo_op": descricao_insumo_op,
						"quant_utilizada_op":quant_utilizada_op,
						"quant_produzida":quant_produzida,
						"quant_total_utilizada":quant_total_utilizada,
						"ult_compra_custo_utilizado_op":ult_compra_custo_utilizado_op, 
						"comentario_ultima_compra_op":comentario_ultima_compra_op,
						"fechamento_custo_utilizado_op":fechamento_custo_utilizado_op, 
						"comentario_fechamento_op":comentario_fechamento_op,
						"medio_atual_custo_utilizado_op":medio_atual_custo_utilizado_op,
						"comentario_custo_medio_op":comentario_custo_medio_op,
						"emissao_op":emissao_op,
						"fechamento_op": fechamento_op,
					})

		op = None
		insumo_op = None
		descricao_insumo_op = None
		quant_utilizada_op = None
		quant_produzida = None
		quant_total_utilizada = None
		ult_compra_custo_utilizado_op = None
		comentario_ultima_compra_op = None
		fechamento_custo_utilizado_op = None
		comentario_fechamento_op = None
		medio_atual_custo_utilizado_op = None
		comentario_custo_medio_op = None
		emissao_op = None
		fechamento_op = None
	else:
		op = consulta_filtrada["op"].iloc[0]
		insumo_op = consulta_filtrada["insumo"].iloc[0]
		descricao_insumo_op = consulta_filtrada["descricao_insumo"].iloc[0]
		quant_utilizada_op = consulta_filtrada["quant_utilizada"].iloc[0]
		quant_produzida = consulta_filtrada["quant_produzida"].iloc[0]
		quant_total_utilizada = consulta_filtrada["quant_total_utilizada"].iloc[0]
		ult_compra_custo_utilizado_op = consulta_filtrada["ult_compra_custo_utilizado"].iloc[0]
		comentario_ultima_compra_op = consulta_filtrada["comentario_ultima_compra"].iloc[0]
		fechamento_custo_utilizado_op = consulta_filtrada["fechamento_custo_utilizado"].iloc[0]
		comentario_fechamento_op = consulta_filtrada["comentario_fechamento"].iloc[0]
		medio_atual_custo_utilizado_op = consulta_filtrada["medio_atual_custo_utilizado"].iloc[0]
		comentario_custo_medio_op = consulta_filtrada["comentario_custo_medio"].iloc[0]
		emissao_op = consulta_filtrada["data_referencia"].iloc[0]
		fechamento_op = consulta_filtrada["data_encerramento_op"].iloc[0]

	return pd.Series({
		"op":op, 
		"insumo_op":insumo_op,
		"descricao_insumo_op":descricao_insumo_op,
		"quant_utilizada_op":quant_utilizada_op,
		"quant_produzida":quant_produzida,
		"quant_total_utilizada":quant_total_utilizada,
		"ult_compra_custo_utilizado_op":ult_compra_custo_utilizado_op, 
		"comentario_ultima_compra_op":comentario_ultima_compra_op,
		"fechamento_custo_utilizado_op":fechamento_custo_utilizado_op, 
		"comentario_fechamento_op":comentario_fechamento_op,
		"medio_atual_custo_utilizado_op":medio_atual_custo_utilizado_op,
		"comentario_custo_medio_op":comentario_custo_medio_op,
		"emissao_op": emissao_op,
		"fechamento_op": fechamento_op,
	})




def combina_estrutura_e_op(estrutura:pd.DataFrame, consulta_op:pd.DataFrame):
	estrutura[[
		"op",
		"insumo_op",
		"descricao_insumo_op",
		"quant_utilizada_op",
		"quant_produzida",
		"quant_total_utilizada",
		"ult_compra_custo_utilizado_op", 
		"comentario_ultima_compra_op",
		"fechamento_custo_utilizado_op", 
		"comentario_fechamento_op",
		"medio_atual_custo_utilizado_op",
		"comentario_custo_medio_op",
		"emissao_op",
		"fechamento_op",
	]] = estrutura.apply(linha_a_linha_estru_op, axis=1, args=(consulta_op,))
	produtos_nao_utilizados = consulta_op[~consulta_op["insumo"].isin(estrutura["insumo_op"])].rename(columns={
		"insumo":"insumo_op",
		"descricao_insumo":"descricao_insumo_op",
		"quant_utilizada": "quant_utilizada_op",
		"ult_compra_custo_utilizado":"ult_compra_custo_utilizado_op",
		"comentario_ultima_compra": "comentario_ultima_compra_op",
		"fechamento_custo_utilizado": "fechamento_custo_utilizado_op",
		"comentario_fechamento": "comentario_fechamento_op",
		"medio_atual_custo_utilizado": "medio_atual_custo_utilizado_op",
		"comentario_custo_medio": "comentario_custo_medio_op",
		"data_referencia": "emissao_op",
		"data_encerramento_op": "fechamento_op"
	})

	estrutura = pd.concat([estrutura, produtos_nao_utilizados],ignore_index=True).fillna("")

	numero_op = consulta_op.loc[consulta_op["op"].notnull(),"op"].values[0]
	estrutura["op"] = numero_op

	descricao_original = estrutura.loc[estrutura["descricao_cod_original"].notnull(),"descricao_cod_original"].values[0]
	estrutura["descricao_cod_original"] = descricao_original

	#estrutura.to_excel("estrutura.xlsx", "relatorio", engine="openpyxl")
	return estrutura




def combina_custos_totais_estrutura_e_op(custos_totais_estrutura: pd.DataFrame, custos_totais_op):

	custos_totais_estrutura_op = custos_totais_estrutura.merge(
		custos_totais_op,how="left",on=["codigo_original","descricao_cod_original","data_referencia"]
	)
	#custos_totais_estrutura_op.to_excel("custos_totais_estrutura_op.xlsx","a", engine="openpyxl")

	return custos_totais_estrutura_op




def gerar_relatorio_excel_bomxop_simples(estrutura_com_op, custos_totais_estrutura_op, data_referencia):
	data_referencia = tratamento_data_referencia(data_referencia)

	colunas_numericas = {
		"quant_utilizada",
		'ult_compra_custo_utilizado',
		'ult_compra_custo_utilizado_alt',
		'fechamento_custo_utilizado',
		'fechamento_custo_utilizado_alt',
		'medio_atual_custo_utilizado',
		'medio_atual_custo_utilizado_alt',
		"quant_utilizada_op",
		"quant_produzida",
		"quant_total_utilizada",
		'ult_compra_custo_utilizado_op',
		'fechamento_custo_utilizado_op',
		'medio_atual_custo_utilizado_op',
	}

	max_lengths = {}

	for col in estrutura_com_op.columns:
		if col in colunas_numericas:
			max_length = 13
		else:  
			max_length = max(estrutura_com_op[col].astype(str).apply(len).max(), len(str(col)))  
		
		max_lengths[col] = max_length

	nomes_das_colunas = {
		'codigo_original':"Cód Original",
		'descricao_cod_original':"Descrição Orig",
		'codigo_pai':"Código Pai", 
		'descricao_pai':"Descrição Pai", 
		'tipo_pai':"Tipo Pai", 
		'insumo':"Insumo BOM", 
		'descricao_insumo':"Descrição Insumo BOM",
		'quant_utilizada':"Quant Utilizada BOM", 
		'tipo_insumo':"Tipo Insumo BOM", 
		'origem':"Origem", 
		'alternativos':"Alternativos", 
		'ult_compra_custo_utilizado':"U Compra BOM",
		'ult_compra_custo_utilizado_alt':"U Compra Alt BOM",
		'fechamento_custo_utilizado':"U Fech BOM",
		'fechamento_custo_utilizado_alt':"U Fecham Alt BOM",
		'medio_atual_custo_utilizado':"Médio Atual BOM",
		'medio_atual_custo_utilizado_alt':"Méd Atu Alt BOM",
		"op":"OP",
		"insumo_op": "Insumo OP",
		"descricao_insumo_op": "Descrição Insumo OP",
		"quant_utilizada_op": "Qtd Util OP (UN)",
		"quant_produzida": "Qtd Prod OP",
		"quant_total_utilizada": "Qtd Total Util OP",
		'ult_compra_custo_utilizado_op':"U Compra OP",
		'fechamento_custo_utilizado_op':"U Fech OP",
		'medio_atual_custo_utilizado_op':"Médio Atual OP",
		"emissao_op": "Data Emissão OP",
		"fechamento_op": "Data Fech OP",
	}

	wb = Workbook()
	ws = wb.active
	ws.title = "Detalhamento BOM x OP"

	ws.append([
		nomes_das_colunas["codigo_original"],
		nomes_das_colunas["descricao_cod_original"],
		nomes_das_colunas["codigo_pai"],
		nomes_das_colunas["descricao_pai"],
		nomes_das_colunas["tipo_pai"],
		nomes_das_colunas["insumo"],
		nomes_das_colunas["descricao_insumo"],
		nomes_das_colunas["quant_utilizada"],
		nomes_das_colunas["tipo_insumo"],
		nomes_das_colunas["origem"],
		nomes_das_colunas["alternativos"],
		nomes_das_colunas["ult_compra_custo_utilizado"],
		nomes_das_colunas["ult_compra_custo_utilizado_alt"],
		# "U Compra BOM TabDin",
		nomes_das_colunas["fechamento_custo_utilizado"],
		nomes_das_colunas["fechamento_custo_utilizado_alt"],
		# "U Fech BOM TabDin",
		nomes_das_colunas["medio_atual_custo_utilizado"],
		nomes_das_colunas["medio_atual_custo_utilizado_alt"],
		# "C Medio BOM TabDin",
		nomes_das_colunas["op"],
		nomes_das_colunas["insumo_op"],
		nomes_das_colunas["descricao_insumo_op"],
		nomes_das_colunas["quant_utilizada_op"],
		nomes_das_colunas["quant_produzida"],
		nomes_das_colunas["quant_total_utilizada"],
		nomes_das_colunas['ult_compra_custo_utilizado_op'],
		nomes_das_colunas['fechamento_custo_utilizado_op'],
		nomes_das_colunas['medio_atual_custo_utilizado_op'],
		nomes_das_colunas["emissao_op"],
		nomes_das_colunas["fechamento_op"],
	])

	for i, row in estrutura_com_op.iterrows():
		l = i+2
		ws.cell(l, 1, row["codigo_original"])
		ws.cell(l, 2, row["descricao_cod_original"])
		ws.cell(l, 3, row["codigo_pai"])
		ws.cell(l, 4, row["descricao_pai"])
		ws.cell(l, 5, row["tipo_pai"])
		ws.cell(l, 6, row["insumo"])
		ws.cell(l, 7, row["descricao_insumo"])
		ws.cell(l, 8, row["quant_utilizada"])
		ws.cell(l, 9, row["tipo_insumo"])
		ws.cell(l,10, row["origem"])
		ws.cell(l,11, row["alternativos"])

		comm = row["comentario_ultima_compra"]
		ws.cell(l,12, row["ult_compra_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_ultima_compra_alt"]
		# ws.cell(l,13).number_format = "@"
		ws.cell(l,13, row["ult_compra_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_fechamento"]
		ws.cell(l,14, row["fechamento_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_fechamento_alt"]
		# ws.cell(l,15).number_format = "@"
		ws.cell(l,15, row["fechamento_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio"]
		ws.cell(l,16, row["medio_atual_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_custo_medio_alt"]
		# ws.cell(l,17).number_format = "@"
		ws.cell(l,17, row["medio_atual_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		ws.cell(l,18, row["op"])
		ws.cell(l,19, row["insumo_op"])
		ws.cell(l,20, row["descricao_insumo_op"])
		ws.cell(l,21, row["quant_utilizada_op"])
		ws.cell(l,22, row["quant_produzida"])
		ws.cell(l,23, row["quant_total_utilizada"])

		comm = row["comentario_ultima_compra_op"]
		ws.cell(l,24).number_format = "@"
		ws.cell(l,24, row['ult_compra_custo_utilizado_op'])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_fechamento_op"]
		ws.cell(l,25).number_format = "@"
		ws.cell(l,25, row['fechamento_custo_utilizado_op'])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio_op"]
		ws.cell(l,26).number_format = "@"
		ws.cell(l,26, row['medio_atual_custo_utilizado_op'])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		ws.cell(l,27, row["emissao_op"]).number_format = "DD/MM/YYYY"
		ws.cell(l,28, row["fechamento_op"]).number_format = "DD/MM/YYYY"


	colunas_para_ocultar = {
		"Descrição Orig","Descrição Pai","Descrição Insumo BOM", "Descrição Insumo OP"
	}

	for col in range(1, ws.max_column + 1):
		cabecalho = ws.cell(1,col).value
		letra_coluna = get_column_letter(col)

		if cabecalho in colunas_para_ocultar:
			ws.column_dimensions[letra_coluna].outlineLevel = 1
			ws.column_dimensions[letra_coluna].hidden = True
		
		for chave, valor in nomes_das_colunas.items():
			if valor == cabecalho:
				adjusted_width = max_lengths[chave] + 4
				ws.column_dimensions[letra_coluna].width = adjusted_width
				break

	ult_coluna = get_column_letter(ws.max_column)
	tab = Table(
		displayName="tabela_estruturas",
		ref=f"A1:{ult_coluna}{len(estrutura_com_op) + 1}"
	)
	style = TableStyleInfo(
		name="TableStyleMedium2",
		showFirstColumn=False,
		showLastColumn=False,
		showRowStripes=True,
		showColumnStripes=False,
	)
	tab.tableStyleInfo = style
	ws.add_table(tab)

	##### CUSTOS_TOTAIS
	ws2 = wb.create_sheet("Consolidado",0)
	ws2.append([
		"OP",
		"Data de Referência", "Código","Descrição",
		"U Entradas BOM",
		"U Entradas OP",
		"U Fechamento BOM",
		"U Fechamento OP",
		"Custo Médio BOM",
		"Custo Médio OP"
	])

	for i, row in custos_totais_estrutura_op.iterrows():
		l = i+2
		ws2.cell(l, 1, row["op"])
		ws2.cell(l, 2, data_referencia).number_format = "DD/MM/YYYY"
		ws2.cell(l, 3, row["codigo_original"])
		ws2.cell(l, 4, row["descricao_cod_original"])

		comm = row["comentario_ultima_compra"]
		ws2.cell(l, 5, row["custo_total_ultima_compra"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_ultima_compra_op"]
		ws2.cell(l, 6, row["custo_total_ultima_compra_op"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_ultimo_fechamento"]
		ws2.cell(l, 7, row["custo_total_ultimo_fechamento"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_fechamento_op"]
		ws2.cell(l, 8, row["custo_total_ult_fechamento_op"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio"]
		ws2.cell(l, 9, row["total_pelo_custo_medio"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio_op"]
		ws2.cell(l, 10, row["total_pelo_custo_medio_op"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

	ult_coluna = get_column_letter(ws2.max_column)

	tab = Table(
		displayName="tabela_consolidada",
		ref=f"A1:{ult_coluna}{len(custos_totais_estrutura_op) + 1}"
	)
	tab.tableStyleInfo = style
	ws2.add_table(tab)

	caminho_arquivo = path.join(settings.MEDIA_ROOT,"BOM x OP.xlsx")

	wb.save(caminho_arquivo)

	return caminho_arquivo


def get_numeros_OPs_por_periodo(data_inicial, data_final, engine = None):

	if not engine:
		engine = get_engine()

	query = get_query_numeros_op_por_periodo()
	resultado = pd.read_sql(text(query), engine, params={
		"data_inicial": data_inicial,
		"data_final": data_final,
	})

	return resultado["op"].to_list()

def get_numero_op_pelo_produto(produto, engine = None):

	if not engine:
		engine = get_engine()
	
	query = get_query_busca_op_pelo_produto()
	resultado = pd.read_sql(text(query), engine, params={"produto": produto})
	numero_op = str(resultado["op"].values[0])
	return numero_op