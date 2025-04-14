from django.shortcuts import render,redirect
from django.contrib import messages
from django.conf import settings
from django.urls import reverse
from django.http import FileResponse

from pandas import read_sql, DataFrame, concat, read_csv
from sqlalchemy import text
from re import search
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter as gcl
from openpyxl.styles import Font,Border,Side,Color
from openpyxl.comments import Comment
from os import path
import xlwings as xw

from ghost.utils.funcs import (
	tratamento_data_referencia, forma_string_para_query, get_cabecalhos_e_rows_dataframe,
	rgb_para_hex
)
from ghost.queries import (
	get_query_numeros_op_varios_periodos_somente_PA
)
from ghost.utils.funcs import get_engine, get_info_produtos
from ghost.views.explop import explode_estrutura_pela_op
from ghost.models import Processamento

def bomxop_linha_do_tempo(request):
	return render(request, "ghost/BOMxOP/bomxop_linha.html", {"caller":"bomxop_linha"})

def bomxop_linha_do_tempo_post(request):

	if request.method != "POST":
		return render(request, "ghost/BOMxOP/bomxop_linha.html")

	qual_custo = request.POST.get("qual-custo")

	datas_iniciais_str = request.POST.getlist("data-inicial")
	datas_iniciais = []
	datas_finais_str = request.POST.getlist("data-final")
	datas_finais = []
	if len(datas_iniciais_str) == 0 or len(datas_finais_str) == 0:
		messages.info(request, "Datas Inválidas")
		return render(request, "ghost/BOMxOP/bomxop_linha.html")
	
	################ processamento
	codigo_identificador = request.POST.get("codigo-identificador","")
	processamento = Processamento.objects.filter(codigo_identificador = codigo_identificador)
	if processamento.exists():
		processamento.delete()

	processamento = Processamento.objects.create(
		codigo_identificador = codigo_identificador,
		caller = "bomxop_linha",
		porcentagem = "0%",
		mensagem1 = "Processando",
		mensagem2 = "Preparando informações"
	)
	################
	
	for data_inicial in datas_iniciais_str:
		if data_inicial:
			data_inicial = tratamento_data_referencia(data_inicial)
			datas_iniciais.append(data_inicial)
	for data_final in datas_finais_str:
		if data_final:
			data_final = tratamento_data_referencia(data_final)
			datas_finais.append(data_final)

	if len(datas_iniciais) != len(datas_finais):
		messages.info(request, "Erro no processamento das datas digitadas")
		return render(request, "ghost/BOMxOP/bomxop_linha.html")
	
	traz_prod = request.POST.get("traz-produzidos")
	
	if traz_prod != "on":
		codigos_pre = request.POST.get("codigos")
		if not codigos_pre:
			messages.info(request, "Códigos Inválidos")
			return render(request, "ghost/BOMxOP/bomxop_linha.html")
		codigos_pre = codigos_pre.split("\r\n")
		codigos = forma_string_para_query(list({str(c).upper().strip() for c in codigos_pre if c}))
	else:
		codigos = None
	
	engine = get_engine()

	traz_budget = True if request.POST.get("traz-budget") == "on" else False
	titulo_budget = "BUDGET25"
	
	if traz_budget:

		################ processamento
		processamento.porcentagem = '1.0%'
		processamento.mensagem2 = "Buscando budget"
		processamento.save()
		################
		
		budget = read_csv("bases/budget25.csv",delimiter=";",decimal=",")

		info_prod = get_info_produtos(forma_string_para_query(budget["codigo_original"].drop_duplicates()), engine)
		budget = budget.merge(info_prod,how="left",left_on="codigo_original",right_on="codigo").drop(columns=["origem_produto","codigo"])\
			.rename(columns={"descricao_produto":"descricao_cod_original","tipo_produto":"tipo_original"})

		info_prod = get_info_produtos(forma_string_para_query(budget["codigo_pai"].drop_duplicates()), engine)
		budget = budget.merge(info_prod,how="left",left_on="codigo_pai",right_on="codigo").drop(columns=["origem_produto","codigo"])\
			.rename(columns={"descricao_produto":"descricao_pai","tipo_produto":"tipo_pai"})
		
		info_prod = get_info_produtos(forma_string_para_query(budget["insumo"].drop_duplicates()),engine)
		budget = budget.merge(info_prod,how="left",left_on="insumo",right_on="codigo").drop(columns=["origem_produto","codigo"])\
			.rename(columns={"descricao_produto":"descricao_insumo","tipo_produto":"tipo_insumo"})

		budget["op_original"] = "00000000000"
		budget["data_encerramento_op_original"] = titulo_budget

		if qual_custo == "uf":
			budget = budget.rename(columns={"ult_compra_custo_utilizado":"fechamento_custo_utilizado"})
		#date(2024,10,19)


	################ processamento
	processamento.porcentagem = '2.0%'
	processamento.mensagem2 = "Buscando OPs do período"
	processamento.save()
	################


	query_ops = get_query_numeros_op_varios_periodos_somente_PA(datas_iniciais, datas_finais)
	ops = read_sql(text(query_ops), engine, params={
		"codigos": codigos,
		"tipos_apontamento": '200'
	})
	query_ops = None
	data_final = None
	datas_finais = None
	datas_finais_str = None
	data_inicial = None
	datas_iniciais = None
	datas_iniciais_str = None

	if ops.empty:
		messages.info(request, "Sem OPs finalizadas no período")
		return render(request, "ghost/BOMxOP/bomxop_linha.html")

	detal_ops = DataFrame()

	################ processamento
	processamento.porcentagem = '3.0%'
	processamento.mensagem1 = "Obtendo estruturas por OPs"
	processamento.save()
	################

	quant_ops = ops.shape[0]
	
	for i, row in ops.iterrows():

		################ processamento
		processamento.mensagem2 = f"{row['produto']} - {row['descricao']}"
		processamento.save()
		################
		
		op = row["op"]
		data = row["data_encerramento_op"]
		request, df_op, coluna_custo = explode_estrutura_pela_op(
			request=request, 
			op=op, 
			engine=engine, 
			data=data, 
			explodir_pis=True,
			qual_custo=qual_custo,
			processamento_dict={
				"processamento":processamento,
				"teto": round((i+1)/quant_ops*87,0)+3
			}
		)
		df_op = df_op.rename(columns={
			"codigo_original":"codigo_pai",
			"descricao_cod_original": "descricao_pai",
			"tipo_original": "tipo_pai"
		})
		df_op["ordem"] = i + 1
		df_op["op_original"] = op
		df_op["data_encerramento_op_original"] = data
		df_op["codigo_original"] = row["produto"]
		df_op["descricao_cod_original"] = row["descricao"]
		df_op["tipo_original"] = row["tipo"]

		detal_ops = concat([detal_ops,df_op], ignore_index=True)

		################ processamento
		processamento.porcentagem = f"{round((i+1)/quant_ops*87,0)+3}%"
		processamento.save()
		################
		
	
	detal_ops = detal_ops.drop(columns=["quant_total_utilizada"])


	if traz_budget:
		budget = budget.loc[budget["codigo_original"].isin(ops["produto"]),:]
		detal_ops = concat([budget,detal_ops]).reset_index()


	################ processamento
	processamento.mensagem2 = f"Separando commodities"
	processamento.save()
	################

	detal_ops["commodity"] = detal_ops.apply(commodity,axis=1)
	
	if qual_custo == "uesf":
		nova_col_custo = "ult_compra_sem_frete"
	elif qual_custo == "uecf":
		nova_col_custo = "ult_compra_com_frete"
	elif qual_custo == "uf":
		nova_col_custo = "ult_fechamento"
	detal_ops = detal_ops.rename(columns={coluna_custo:nova_col_custo})
	
	################ processamento
	processamento.mensagem1 = "Preparando informações para tela"
	processamento.save()
	################

	dicio_tela = {}
	colunas_somadas_resumo = {}
	codigos_originais = list(detal_ops["codigo_original"].drop_duplicates())
	quantos_codigos = len(codigos_originais)
	contagem_codigos = 0

	for prod in codigos_originais:
		
		descricao_cod_original = detal_ops.loc[detal_ops["codigo_original"] == prod,"descricao_cod_original"].iloc[0]
		prod_completo = str(prod) + " " + str(descricao_cod_original)

		################ processamento
		processamento.mensagem2 = f"{prod_completo}"
		processamento.save()
		################

		dicio_tela.update({prod_completo:{}})
		colunas_somadas_resumo.update({prod_completo:{}})
		soma_por_coluna = {}
		for commod in detal_ops.loc[detal_ops["codigo_original"] == prod,"commodity"].drop_duplicates():
			dicio_tela[prod_completo].update({commod:[]})
			colunas_somadas_resumo[prod_completo].update({commod:[]})

			temp_df = detal_ops.loc[
				(detal_ops["codigo_original"] == prod) & 
				(detal_ops["commodity"] == commod), :
			].pivot_table(
				values=[nova_col_custo],
				index=["codigo_original","descricao_cod_original","tipo_original",
				"commodity","insumo","descricao_insumo","tipo_insumo"],
				columns=["op_original","data_encerramento_op_original"],
				aggfunc="sum"
			).reset_index()

			temp_df.columns = temp_df.columns.map(lambda x: '<br>'.join(str(i) for i in x))

			temp_df = temp_df.fillna(0)

			temp_df = temp_df.drop(columns=["codigo_original<br><br>","descricao_cod_original<br><br>","tipo_original<br><br>"])

			colunas_somadas = []
			for col in temp_df.columns:
				if nova_col_custo in col:
					soma = round(temp_df[col].sum(),5)
					temp_df[col] = round(temp_df[col],5)
					colunas_somadas.append(soma)
					colunas_somadas_resumo[prod_completo][commod].append((col.split("<br>")[2],soma))

					if col not in soma_por_coluna:
						soma_por_coluna.update({col:soma})
					else:
						soma_por_coluna[col] += soma

				else:
					colunas_somadas.append("")

			cabecalhos, rows = get_cabecalhos_e_rows_dataframe(temp_df)
				
			dicio_tela[prod_completo][commod].append([
				cabecalhos,
				colunas_somadas,
				rows
			])

		colunas_somadas_resumo[prod_completo].update({"TOTAL":[]})
		for key,val in soma_por_coluna.items():
			colunas_somadas_resumo[prod_completo]["TOTAL"].append((key.split("<br>")[2],round(val,5)))

		################ processamento
		processamento.porcentagem = f"{round(contagem_codigos/quantos_codigos*8,0)+90}%"
		contagem_codigos += 1
		processamento.save()
		################


	################ processamento
	processamento.porcentagem = "99%"
	processamento.mensagem1 = "Preparando relatório em excel"
	processamento.mensagem2 = ""
	processamento.save()
	################

	############### RELATORIO EXCEL
	wb = Workbook()
	ws = wb.worksheets[0]
	ws.title = "Relatório"

	prim_lin = wslin = 3
	prim_col = wscol = 1

	texto_na_direita = Alignment(horizontal="right")
	texto_na_direita_br = Alignment(horizontal="right",wrap_text=True)
	negrito = Font(bold=True)
	borda_pontilhada_style = Side(style="dashDot")
	borda_solida_style = Side(style="thin")
	borda_pontilhada_preta = Border(
		left=borda_pontilhada_style,
		right=borda_pontilhada_style,
		top=borda_pontilhada_style,
		bottom=borda_pontilhada_style
	)
	borda_solida_inferior = Border(bottom=borda_solida_style)
	vermelho_escuro_cor = Color(rgb_para_hex(156,0,6))
	verde_escuro_cor = Color(rgb_para_hex(0,97,0))
	fonte_vermelha = Font(color=vermelho_escuro_cor,bold=True)
	fonte_verde = Font(color=verde_escuro_cor,bold=True)
	fonte_grande = Font(size=16)


	outline_1_lins = set()
	deltas = {}

	ws.column_dimensions[gcl(wscol)].width = 50
	ws.cell(1,1,"OPs em Linha do Tempo").font = fonte_grande
	############### RELATORIO EXCEL

	for prod, prod_val in colunas_somadas_resumo.items():
		wscol = prim_col
		celula = ws.cell(wslin,wscol,prod)
		celula.border = borda_solida_inferior
		lin_prod = wslin
		wslin += 1
		prim_lin_outline = wslin

		for commod, commod_val in prod_val.items():
			wscol = prim_col
			if commod == "TOTAL": wslin += 1
			celula = ws.cell(wslin,wscol,commod)
			celula.alignment = texto_na_direita
			if commod == "TOTAL": celula.font = negrito
			
			wscol += 1
			
			lin_commod = wslin
			for cc, vv in commod_val:
				wslin = lin_commod
				celula = ws.cell(wslin,wscol,cc)
				celula.alignment = texto_na_direita
				if commod == "TOTAL": celula.font = negrito
				wslin += 1
				celula = ws.cell(wslin,wscol,vv)
				if commod == "TOTAL": celula.font = negrito

				if traz_budget:
					if cc == titulo_budget:
						valor_budget = vv
					valor_ult_op = vv
					if valor_ult_op > valor_budget + 0.01:
						celula.font = fonte_vermelha
					elif valor_ult_op < valor_budget - 0.01:
						celula.font = fonte_verde
				
				wslin += 1
				wscol += 1
			
			if traz_budget:
				delta = round(valor_budget-valor_ult_op,2)
				celula = ws.cell(lin_prod,prim_col + 1,delta)
				celula.font = fonte_verde if delta >= 0 else fonte_vermelha
				celula.border = borda_solida_inferior
				celula = ws.cell(lin_prod,prim_col + 2,"\u2206")
				celula.font = fonte_verde if delta >= 0 else fonte_vermelha
				deltas.update({prod:delta})
				

			if commod != "TOTAL":
				for cabecalhos, colunas_somadas, rows in dicio_tela[prod][commod]:
					wscol = prim_col + 1
					for val in cabecalhos:
						val = str(val).replace("<br>","\n")
						celula = ws.cell(wslin,wscol,val)
						celula.alignment = texto_na_direita_br
						celula.border = borda_pontilhada_preta
						ws.column_dimensions[gcl(wscol)].width = 22
						wscol += 1
					outline_1_lins.add(wslin)
					wslin += 1
					wscol = prim_col + 1
					for val in colunas_somadas:
						celula = ws.cell(wslin,wscol,val)
						celula.border = borda_pontilhada_preta
						wscol += 1
					outline_1_lins.add(wslin)
					wslin += 1
					for row in rows:
						wscol = prim_col + 1
						for key,val in row.items():
							if titulo_budget in key and traz_budget:
								val_budget = val
							celula = ws.cell(wslin,wscol,val)
							if type(val) == float and traz_budget:
								if val > val_budget + 0.01:
									celula.font = fonte_vermelha
								elif val < val_budget - 0.01:
									celula.font = fonte_verde
							celula.border = borda_pontilhada_preta
							wscol += 1
						outline_1_lins.add(wslin)
						wslin += 1

		for i in range(prim_lin_outline,wslin-2):
			ws.row_dimensions[i].outlineLevel = 1
			ws.row_dimensions[i].hidden = True
		for i in outline_1_lins:
			ws.row_dimensions[i].outlineLevel = 2
			ws.row_dimensions[i].hidden = True
		
		wslin += 1
	
	if traz_budget:
		deltas_decrescente = "\n".join(f"{k}: {v}" for k,v in sorted(deltas.items(),key=lambda x:x[1],reverse=False))
		ws.cell(1, 1).comment = \
			Comment(
				deltas_decrescente, "", 
				(deltas_decrescente.count("\n") + 2) * 20, 
				max(len(lin) for lin in deltas_decrescente.split("\n")) * 8
			)

		
	ws.sheet_properties.outlinePr.summaryBelow = True 
	ws.sheet_view.showGridLines = False
		
	caminho_arquivo = path.join(settings.MEDIA_ROOT,"OP em Linha do Tempo.xlsx")
	wb.save(caminho_arquivo)

	app = xw.App(visible=False,add_book=False)
	xwwb = app.books.open(caminho_arquivo)
	xwws = xwwb.sheets("Relatório")
	xwws.api.Outline.ShowLevels(RowLevels=2)
	xwws.api.Outline.ShowLevels(RowLevels=1)
	xwwb.save()
	xwwb.close()
	app.quit()

	################ processamento
	processamento.porcentagem = "100%"
	processamento.mensagem1 = "Concluído"
	processamento.save()
	################

	context = {
		"caller": "bomxop_linha_post",
		"dicio_tela": dicio_tela,
		"colunas_somadas_resumo":colunas_somadas_resumo,
		"caminho_arquivo":caminho_arquivo,
	}

	return render(request, "ghost/BOMxOP/bomxop_linha_post.html", context)




def commodity(row):

	if row["tipo_pai"] == "PI":
		descri_pai = row["descricao_pai"]
		if search(r"(?:^| )(HOUSING|INJ|COVER|HOLDER|CARCACA)(?:$| )",descri_pai):
			return "INJECAO"
		elif search(r"(?:^| )(CALEFATOR|JUMPER|TRIAC|DIOD[OE]|HEATER|HELICAL|ION WIRE CLAMP)(?:$| )",descri_pai):
			return "CALEFATOR"
		elif search(r"(?:^| )(CONTACT SET|CHAPA CONTATO|CONTACT PLATE|SENSOR ELEG)(?:$| )",descri_pai):
			return "PATIN"
		elif search(r"(?:^| )(GRID|FILTER|GRADE)(?:$| )", descri_pai):
			return "ESTAMPARIA"
		elif search(r"(?:^| )(PCB[A]?)(?:$| )", descri_pai):
			return "PLACA"

	if row["tipo_insumo"] == "EM":
		return "EMBALAGEM"
	elif row["tipo_insumo"] == "BN":
		return "SERV. INDUST. TERCEIRIZADA"

	texto = str(row["descricao_insumo"]).upper().strip()

	if search(r"SWITCH",texto):
		return "CHAVES"
	elif search(r"^(MOTOR|SET MOTOR)", texto):
		return "MOTOR"
	elif search(r"(PC HS|PC HD|POWER CORD)", texto):
		return "CABO DE FORCA"
	elif search(r"^(PA[ ]?66|PC |MASTER PC)", texto):
		return "INJECAO"
	elif search(r"(^TINTA |TAMPOGRAFIA| TAMPO |CATALI[ZS]ADOR)",texto):
		return "TAMPOG E PINTURA"
	elif search(r"(?:^| )(MICA|RIVET)(?:$| )", texto):
		return "CALEFATOR"
	elif search(r"(ALUMIN[I]?UM PLATE|T[H]?ERMISTOR PTC)", texto):
		return "PATIN"
	elif search(r"(?:^| )(PCB[A]?)(?:$| )", texto):
		return "PLACA"
	elif search(r"(?:^| )(PLAST[I]?[C]?)(?:$| )", texto):
		return "PECAS PLASTICAS"
	elif search(r"(?:^| )(HEATER)(?:$| )", texto):
		return "CALEFATOR"
	elif search(r"(?:^| )(HOUSING)", texto):
		return "HOUSING"
	elif search(r"(?:^| )(DIFFUSER)(?:$| )", texto):
		return "DIFUSOR"
	elif search(r"(?:^| )(RUBBER PAD)(?:$| )", texto):
		return "RUBBER PAD"
	elif search(r"(?:^| )(FILTER)(?:$| )", texto):
		return "FILTER"
	
	else:
		return "OUTROS"




def baixar_relatorio_bomxop_linha(request):
	if request.method != "POST": return redirect(reverse("ghost:bomxop-linha-do-tempo"))

	caminho_arquivo = request.POST.get("caminho-arquivo")
	
	if not caminho_arquivo:
		messages.error(request, "Algo saiu errado na geração do relatório.")
		return redirect(reverse("ghost:bomxop-linha-do-tempo"))

	return FileResponse(open(caminho_arquivo,"rb"),as_attachment=True, filename="OPs em Linha do Tempo.xlsx")
