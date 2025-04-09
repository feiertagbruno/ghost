from django.shortcuts import render,redirect
from django.urls import reverse
from django.contrib import messages
from django.conf import settings
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.http import HttpResponse

import pandas as pd
import sqlite3
from sqlalchemy import text
from datetime import datetime
import locale
from dateutil.relativedelta import relativedelta
from io import StringIO
from numpy import nan, where
import json
import xlwings as xw
from win32com.client import constants
from os import path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font,Color, Border,Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.formatting.rule import CellIsRule

from ghost.queries import get_query_estoque_atual
from ghost.views.estruturas import explode_estrutura, forma_string_codigos, gerar_multiestruturas
from ghost.utils.funcs import (
	get_engine, tratamento_data_referencia, gerar_codigo_aleatorio_simulador,
	get_info_produtos, rgb_para_long, rgb_para_hex
)
from ghost.views.consultas import get_produzidos_na_data,get_pedidos
from ghost.models import Processamento

def simulador_de_producao(request):
	if request.session.get("codigo-aleatorio"):
		del(request.session["codigo-aleatorio"])
	
	query_tabelas_salvas = "SELECT name FROM sqlite_master WHERE type = 'table' AND name like 'simulacao_%'"
	sqlite_conn = sqlite3.connect("db.sqlite3")

	tabelas = []
	tabelas_existentes = pd.read_sql(query_tabelas_salvas, sqlite_conn)
	if not tabelas_existentes.empty:
		lista_tabelas = tabelas_existentes["name"].to_list()

		for tb in lista_tabelas:
			tabelas.append(tb.split("_",1)[1])

	context = {
		"caller":"inicial",
		"tabelas":tabelas if tabelas else None,
	}
	return render(request, "ghost/simuladordeproducao/simuladordeproducao.html", context)


def adicionar_producao(request):
	if not request.method == "POST":
		return redirect(reverse("ghost:simulador-de-producao"))
	
	tabela_salva = request.POST.get("tabela-salva")

	if tabela_salva:
		return adicionar_nova_producao(request, tabela_salva)
	
	codigo_aleatorio = request.session.get("codigo-aleatorio")

	engine = get_engine()
	
	codigo = request.POST.get("codigo-produto")
	data = request.POST.get("data-producao")
	quant = request.POST.get("quantidade")
	explode_pis = True if request.POST.get("explode-pis", None) else False
	abre_detalhamento = False if request.POST.get("abre-detalhamento", None) else True
	
	try:
		quant = float(quant)
	except:
		messages.error(request, "Quantidade inválida")
		return redirect(reverse("ghost:simulador-de-producao"))
	if data:
		data = tratamento_data_referencia(data)

	if data < datetime.today().date():
		messages.error(request, "A data digitada não pode ser menor do que hoje")
		return redirect(reverse("ghost:simulador-de-producao"))

	estrutura, todos_os_codigos = explode_estrutura(
		codigo=codigo,
		data_referencia = data,
		engine = engine,
		abre_todos_os_PIs=explode_pis,
		solicitante='simulador'
	)

	# PRODUZIDOS
	hoje_date = datetime.today().date()
	if data == hoje_date:
		produzidos = get_produzidos_na_data(
			data_inicial=data,
			codigos=codigo,
			engine=engine
		)
		if not produzidos.empty:
			quant_produzida = produzidos.loc[produzidos["codigo"] == codigo, "quant"].values[0]
			quant -= quant_produzida
	
	if quant <= 0:
		messages.error(request,f"A quantidade produzida neste dia foi de {int(quant_produzida)}. "
				  "Só é possivel simular uma quantidade maior do que esta.")
		return redirect(reverse("ghost:simulador-de-producao"))
		

	estrutura["quant_utilizada"] = estrutura["quant_utilizada"].astype(float).map(lambda x: -round(x * quant,5))
	todos_os_codigos = pd.concat([
		todos_os_codigos, 
		pd.DataFrame({"todos_os_codigos":[codigo]})
	])
	todos_os_codigos = forma_string_codigos(todos_os_codigos["todos_os_codigos"].drop_duplicates())

	query_estoque = get_query_estoque_atual()
	estoque = pd.read_sql(text(query_estoque),engine, params={
		"codigos": todos_os_codigos,
	})

	# FORMA O ESTOQUE_PIVOT
	estoque_pivot = estoque.pivot(index=["codigo","tipo","descricao","origem"], columns="armazem", values="quant").reset_index()
	arm_exist = [col for col in estoque_pivot.columns if len(col) == 2]
	estoque_pivot["Ttl Est"] = estoque_pivot[arm_exist].sum(axis=1)
	hoje = hoje_date.strftime('%d-%m-%Y')
	
	# ESTOQUE NOVOS CABEÇALHOS
	estoque_pivot = padronizar_cabecalhos_estoque(estoque_pivot, hoje)
	
	# ESTRUTURA NOVOS CABEÇALHOS
	data_str = data.strftime('%d-%m-%Y')
	estrutura, coluna_insumo = padronizar_cabecalhos_estrutura(codigo,data_str,quant,estrutura)

	# COMBINA O ESTOQUE_PIVOT COM O ESTRUTURA
	simulador = estoque_pivot.merge(
		estrutura,how="outer",
		left_on=f'Estoquexxx{hoje}xxxcodigo',
		right_on=coluna_insumo
	)
	estoque_pivot = None

	# RESULTADO DO PA/PI PRODUZIDO
	simulador = resultado_pi_pa_produzido(simulador,hoje,codigo,data_str,quant)

	# PREENCHER PRODUTOS QUE VIERAM SOMENTE NA BOM
	simulador = preencher_produtos_sem_estoque(
		simulador, codigo, data_str, quant, hoje
	)

	# PEDIDOS
	pedidos_pivot = get_pedidos_pivot(
		engine=engine,
		coluna_codigos=simulador[f"Estoquexxx{hoje}xxxcodigo"]
	)

	# PEDIDOS NOVOS CABEÇALHOS
	pedidos_pivot = padronizar_cabecalhos_pedidos(pedidos_pivot)

	# PEDIDOS MERGE
	simulador = simulador.merge(
		right=pedidos_pivot,how="left",
		left_on=f"Estoquexxx{hoje}xxxcodigo",
		right_on="Pedidosxxxcodigoxxxquant_pedidos"
	).drop(columns=["Pedidosxxxcodigoxxxquant_pedidos"])

	colunas_ordenadas, datas_encontradas = ordenar_colunas_por_data(simulador.columns)
	simulador = simulador[colunas_ordenadas]
	colunas_ordenadas = None

	# DESCRIÇÃO PARA PRODUTOS SEM DESCRIÇÃO
	simulador = descricao_para_produtos_sem_descricao(simulador,hoje,codigo,data_str,quant,engine)

	# RESULTADO
	simulador.fillna({f"Estoquexxx{hoje}xxxTtl Est":0}, inplace=True)

	simulador_blocos = pd.DataFrame()
	col_resultado_anterior = f"Estoquexxx{hoje}xxxTtl Est"
	for d in datas_encontradas:
		d_str = d.strftime("%d-%m-%Y")
		colunas_para_somar = [col_resultado_anterior]
		bloco_dia = simulador.filter(like=d_str)
		bloco_estoque = bloco_dia.filter(like="Estoque")
		bloco_producao = bloco_dia.filter(like="Produção")
		bloco_pedidos = bloco_dia.filter(like="Pedidos")
		if not bloco_estoque.empty:
			simulador_blocos = pd.concat([simulador_blocos,bloco_estoque], axis=1)
		if not bloco_pedidos.empty:
			colunas_para_somar.append(f"Pedidosxxx{d_str}xxxquant_pedidos")
			simulador_blocos = pd.concat([simulador_blocos,bloco_pedidos], axis=1)
		if not bloco_producao.empty:
			colunas_quant_utilizada = bloco_producao.filter(like="quant_utilizada").columns
			if colunas_quant_utilizada.any():
				colunas_para_somar.extend(colunas_quant_utilizada)
			colunas_quant_utilizada = None
			simulador_blocos = pd.concat([simulador_blocos,bloco_producao], axis=1)


		simulador_blocos[f"Resultadoxxx{d_str}xxxqtd"] = simulador_blocos[colunas_para_somar].sum(axis=1).round(5)
		col_resultado_anterior = f"Resultadoxxx{d_str}xxxqtd"

		bloco_estoque = None
		bloco_producao = None
		bloco_pedidos = None
		colunas_para_somar = []

	simulador = simulador_blocos
	simulador_blocos = None

	# NEGATIVOS MP EM
	simulador = verificar_alternativos_dos_itens_negativos(simulador, data_str, hoje, codigo, quant)
	
	# ADICIONAR EXCLUSIVIDADE
	simulador.insert(0,f"-xxx{hoje}xxxExclusividade","EXCLUSIVO")

	# SALVAR NO BANCO DE DADOS
	request = salvar_dataframe_no_bd(request,simulador,"simudraft",codigo_aleatorio)
	
	simulador = simulador.fillna('')

	# CABEÇALHOS
	cabecalhos, rows = get_cabecalhos_e_rows_simulador_de_producao(
		simulador,hoje, data_str, max(datas_encontradas).strftime("%d-%m-%Y"), abre_detalhamento
	)

	colunas_fixas = get_colunas_fixas(hoje)	
	campos_alteraveis = get_campos_alteraveis()

	context = {
		"caller":"adicionar_producao",
		"cabecalhos":cabecalhos,
		"rows":rows,
		"codigo_aleatorio": codigo_aleatorio,
		"colunas_fixas": colunas_fixas,
		"campos_alteraveis": campos_alteraveis,
		"data_estoque": hoje,
	}

	
	return render(request, "ghost/simuladordeproducao/simuladordeproducao.html",context)

def reprocessar_tabela(request):
	if request.method != "POST":
		return redirect(reverse("ghost:simulador-de-producao"))

	armazens = request.POST.getlist("armazem-checkbox")
	tabela = request.POST.get("tabela-simulador-html")
	if tabela:
		tabela = pd.read_html(StringIO(tabela), header=[0,1,2])[0]

	tb_filtrada = tabela.loc[:, tabela.columns.get_level_values(0) == 'Estoque']
	# tb_filtrada = tabela

	prim_cabecalho = tabela.columns.values[0]

	for col in tabela.columns.values:
		if col[2] == "Ttl Est":
			tabela = tabela.drop(columns=col)

	tabela.loc[:, (prim_cabecalho[0], prim_cabecalho[1], "Ttl Est")] = \
    	tabela.loc[:, tabela.columns.get_level_values(2).isin(armazens)].sum(axis=1)

	cabecalhos = {
		"data": tabela.columns.levels[1],
		"armazem": [*tabela.columns.levels[2]],
		"tam": len(tabela.columns.levels[2]),
	}

	rows = tabela.reset_index(drop=True).to_dict(orient="records")

	context = {
		"cabecalhos": cabecalhos,
		"rows": rows,
		"armazens_visiveis": ['Código','Ttl Est',*armazens],
		"armazens":armazens,
		"caller": "reprocessar_tabela",
	}

	return render(request, "ghost/simuladordeproducao/tabela_simulador.html",context)




@api_view(["POST"])
def altera_simulador_de_producao(request):
	data = request.data
	codigo_aleatorio = data.get("codigo_aleatorio")
	unique = data.get("unique")
	novo_valor = data.get("novo_valor")

	try: 
		novo_valor = float(novo_valor)
	except:
		...

	col_name, row_index = unique.split("|")
	row_index = int(row_index)

	conn = sqlite3.connect("db.sqlite3")
	cursor = conn.cursor()

	query = f"""
UPDATE [{codigo_aleatorio}] 
SET [{col_name}] = '{novo_valor}' 
WHERE [index] = {row_index} 
"""
	try:
		cursor.execute(query)
		conn.commit()
		conn.close()
		response = Response({"sucesso": True})
	except Exception as e:
		response = Response({"erro": e})
		
	return response




def adicionar_nova_producao(request, tabela_salva):

	tabela_salva = f"simulacao_{tabela_salva.strip()}"
	engine = get_engine()
	codigo_aleatorio = request.session.get("codigo-aleatorio")
	codigo = request.POST.get("codigo-produto")
	data = request.POST.get("data-producao")
	quant = request.POST.get("quantidade")
	explode_pis = True if request.POST.get("explode-pis", None) else False
	abre_detalhamento = False if request.POST.get("abre-detalhamento", None) else True

	
	try:
		quant = float(quant)
	except:
		messages.error(request, "Quantidade inválida")
		return redirect(reverse("ghost:simulador-de-producao"))
	if data:
		data = tratamento_data_referencia(data)

	if data < datetime.today().date():
		messages.error(request, "A data digitada não pode ser menor do que hoje")
		return redirect(reverse("ghost:simulador-de-producao"))



	estrutura, todos_os_codigos = explode_estrutura(
		codigo=codigo,
		data_referencia = data,
		engine = engine,
		abre_todos_os_PIs=explode_pis,
		solicitante='simulador'
	)


	# PRODUZIDOS
	hoje_date = datetime.today().date()
	if data == hoje_date:
		produzidos = get_produzidos_na_data(
			data_inicial=data,
			codigos=codigo,
			engine=engine
		)
		if not produzidos.empty:
			quant_produzida = produzidos.loc[produzidos["codigo"] == codigo, "quant"].values[0]
			quant -= quant_produzida
	
	if quant <= 0:
		messages.error(request,f"A quantidade produzida neste dia foi de {int(quant_produzida)}. "
				  "Só é possivel simular uma quantidade maior do que esta.")
		return redirect(reverse("ghost:simulador-de-producao"))
	
	# PEGAR A TABELA SALVA NO BANCO
	sqlite_conn = sqlite3.connect('db.sqlite3')
	simulador = pd.read_sql(f"SELECT * FROM [{tabela_salva}]",sqlite_conn).drop(columns="index")
	sqlite_conn.close()

	data_str = data.strftime('%d-%m-%Y')


	# VERIFICAR SE A PRODUÇÃO JÁ EXISTE NO SIMULADOR
	verif_producao = simulador.filter(like=f"Produçãoxxx{codigo}_{data_str}")
	if not verif_producao.empty:
		simulador.drop(columns=verif_producao.columns, inplace=True)
	verif_producao = None

	# DATA ESTOQUE
	data_estoque = simulador.filter(like="Estoque").filter(like="codigo").columns[0].split("xxx")[1]

	# QUANT_UTILIZADA
	estrutura["quant_utilizada"] = estrutura["quant_utilizada"].astype(float).map(lambda x: -round(x * quant,5))

	# ADICIONAR EXCLUSIVIDADE

	def extrair_produtos_ja_adicionados(simulador):
		produtos_ja_adicionados = set()
		for col in simulador.filter(like="Produçãoxxx").columns:
			produtos_ja_adicionados.add(col.split("xxx")[1].split("_")[0])
		return produtos_ja_adicionados
	
	produtos_ja_adicionados = extrair_produtos_ja_adicionados(simulador)
	if codigo not in produtos_ja_adicionados:
		simulador.loc[simulador[f"Estoquexxx{data_estoque}xxxcodigo"].isin(estrutura["insumo"]),f"-xxx{data_estoque}xxxExclusividade"] = "COMUM"


	todos_os_codigos = pd.concat([
		todos_os_codigos, 
		pd.DataFrame({"todos_os_codigos":[codigo]})
	])
	todos_os_codigos = todos_os_codigos.drop_duplicates("todos_os_codigos")
	todos_os_codigos = todos_os_codigos.loc[
		~todos_os_codigos["todos_os_codigos"].isin(simulador[f"Estoquexxx{data_estoque}xxxcodigo"]),
		"todos_os_codigos"
	]

	# ESTRUTURA NOVOS CABEÇALHOS
	estrutura, coluna_insumo = padronizar_cabecalhos_estrutura(codigo,data_str,quant,estrutura)


	if not todos_os_codigos.empty:

		todos_os_codigos = forma_string_codigos(todos_os_codigos)

		query_estoque = get_query_estoque_atual()
		estoque = pd.read_sql(text(query_estoque),engine, params={
			"codigos": todos_os_codigos,
		})

		
		# FORMA O ESTOQUE_PIVOT
		estoque_pivot = estoque.pivot(index=["codigo","tipo","descricao","origem"], columns="armazem", values="quant").reset_index()
		arm_exist = [col for col in estoque_pivot.columns if len(col) == 2]
		estoque_pivot["Ttl Est"] = estoque_pivot[arm_exist].sum(axis=1)
		
		# ESTOQUE NOVOS CABEÇALHOS
		novos_cabecalhos = {}
		for col in estoque_pivot.columns:
			novos_cabecalhos.update({col:f"Estoquexxx{data_estoque}xxx{col}"})

		estoque_pivot = estoque_pivot.rename(columns=novos_cabecalhos)

		simulador = pd.concat([simulador, estoque_pivot])
	

	# COMBINA O ESTOQUE_PIVOT COM O ESTRUTURA
	simulador = simulador.merge(
		estrutura,how="outer",
		left_on=f'Estoquexxx{data_estoque}xxxcodigo',
		right_on=coluna_insumo
	)

	# RESULTADO DO PA/PI PRODUZIDO
	simulador = resultado_pi_pa_produzido(simulador,data_estoque,codigo,data_str,quant)

	# PREENCHER PRODUTOS QUE VIERAM SOMENTE NA BOM
	simulador = preencher_produtos_sem_estoque(
		simulador, codigo, data_str, quant, data_estoque
	)


	# ORDENAR COLUNAS POR DATA
	colunas_ordenadas, datas_encontradas = ordenar_colunas_por_data(simulador.columns)
	simulador = simulador[colunas_ordenadas]
	

	# DESCRIÇÃO PARA PRODUTOS SEM DESCRIÇÃO
	simulador = descricao_para_produtos_sem_descricao(simulador,data_estoque,codigo,data_str,quant,engine)

	# RESULTADO
	simulador.fillna({f"Estoquexxx{data_estoque}xxxTtl Est":0}, inplace=True)

	simulador_blocos = pd.DataFrame()
	col_resultado_anterior = f"Estoquexxx{data_estoque}xxxTtl Est"
	for d in datas_encontradas:
		d_str = d.strftime("%d-%m-%Y")
		colunas_para_somar = [col_resultado_anterior]
		bloco_dia = simulador.filter(like=d_str)
		bloco_estoque = bloco_dia.filter(like="Estoque")
		bloco_producao = bloco_dia.filter(like="Produção")
		bloco_pedidos = bloco_dia.filter(like="Pedidos")
		if not bloco_estoque.empty:
			simulador_blocos = pd.concat([simulador_blocos,bloco_estoque], axis=1)
		if not bloco_pedidos.empty:
			colunas_para_somar.append(f"Pedidosxxx{d_str}xxxquant_pedidos")
			simulador_blocos = pd.concat([simulador_blocos,bloco_pedidos], axis=1)
		if not bloco_producao.empty:
			colunas_quant_utilizada = bloco_producao.filter(like="quant_utilizada").columns
			if colunas_quant_utilizada.any():
				colunas_para_somar.extend(colunas_quant_utilizada)
			colunas_quant_utilizada = None
			simulador_blocos = pd.concat([simulador_blocos,bloco_producao], axis=1)

		simulador_blocos[f"Resultadoxxx{d_str}xxxqtd"] = simulador_blocos[colunas_para_somar].sum(axis=1).round(5)
		col_resultado_anterior = f"Resultadoxxx{d_str}xxxqtd"

		bloco_estoque = None
		bloco_producao = None
		bloco_pedidos = None
		colunas_para_somar = []


	simulador_blocos.insert(0,f"-xxx{data_estoque}xxxExclusividade",simulador[f"-xxx{data_estoque}xxxExclusividade"])
	simulador = simulador_blocos
	simulador_blocos = None

	# NEGATIVOS MP EM
	simulador = verificar_alternativos_dos_itens_negativos(simulador, data_str, data_estoque, codigo, quant)

	# EXCLUSIVO 2
	simulador.fillna({f"-xxx{data_estoque}xxxExclusividade":"EXCLUSIVO"}, inplace=True)	
	
	# SALVAR NO BANCO DE DADOS
	request = salvar_dataframe_no_bd(
		request=request,
		df=simulador,
		inicial_tabela="simudraft",
		codigo_aleatorio=codigo_aleatorio
	)
	
	simulador = simulador.fillna('')

	# CABEÇALHOS
	cabecalhos, rows = get_cabecalhos_e_rows_simulador_de_producao(
		simulador, data_estoque, data_str, max(datas_encontradas).strftime("%d-%m-%Y"), abre_detalhamento
	)

	colunas_fixas = get_colunas_fixas(data_estoque)
	campos_alteraveis = get_campos_alteraveis()

	context = {
		"caller":"adicionar_nova_producao",
		"cabecalhos":cabecalhos,
		"rows":rows,
		"codigo_aleatorio": codigo_aleatorio,
		"tabela_salva":tabela_salva.split("_",1)[1],
		"colunas_fixas": colunas_fixas,
		"campos_alteraveis": campos_alteraveis,
		"data_str": data_str,
		"data_estoque": data_estoque,
		"data_str_nav": data.strftime("%Y-%m-%d"),
	}

	# simulador.to_excel("a.xlsx")
	return render(request, "ghost/simuladordeproducao/simuladordeproducao.html",context)




def padronizar_cabecalhos_estrutura(codigo:str,data_str:str,quant, estrutura):
	novos_cabecalhos = {}
	
	for col in estrutura.columns:
		if col == 'insumo': coluna_insumo = f"Produçãoxxx{codigo}_{data_str}_{quant}xxx{col}"
		novos_cabecalhos.update({col: f"Produçãoxxx{codigo}_{data_str}_{quant}xxx{col}"})

	estrutura = estrutura.rename(columns=novos_cabecalhos)

	return estrutura, coluna_insumo




def ordenar_colunas_por_data(colunas):

	datas_extraidas = set()

	def extrair_data(col):
		partes = col.split("xxx")
		if partes[0] in ["Estoque","Pedidos","Resultado"]:
			parte=pd.to_datetime(partes[1], format="%d-%m-%Y")
		elif partes[0] == "Produção":
			partes2 = partes[1].split("_")
			parte = pd.to_datetime(partes2[1], format="%d-%m-%Y")
		elif partes[2] == "Exclusividade":
			parte=pd.to_datetime(partes[1], format="%d-%m-%Y")

		datas_extraidas.add(parte.date())
		return parte
	colunas_ordenadas = sorted(colunas, key=extrair_data)
	datas_encontradas = sorted(datas_extraidas)
	
	return colunas_ordenadas, datas_encontradas




@api_view(["POST"])
def salvar_simulacao(request):

	body = request.data
	codigo_aleatorio = body.get("codigo_aleatorio")
	nome_da_simulacao = body.get("nome_da_simulacao")

	nome_da_simulacao = f"simulacao_{nome_da_simulacao.strip()}"

	query_verificacao = f"SELECT name FROM sqlite_master WHERE type='table' AND name='{nome_da_simulacao}'"

	query = f"CREATE TABLE [{nome_da_simulacao}] as SELECT * FROM [{codigo_aleatorio}]"

	try:
		sqlite_conn = sqlite3.connect('db.sqlite3')
		cursor = sqlite_conn.cursor()

		# VERIFICAR SE TABELA EXISTE
		cursor.execute(query_verificacao)
		if cursor.fetchone() is not None:
			cursor.execute(f"DROP TABLE [{nome_da_simulacao}]")

		cursor.execute(query)
		sqlite_conn.commit()
		response = {"sucesso": True}
	except Exception as e:
		response = {"erro":str(e)}

	sqlite_conn.close()

	return Response(response)




def get_colunas_fixas(data_str):
	return [
		f"-xxx{data_str}xxxExclusividade",
		f"Estoquexxx{data_str}xxxcodigo",
		f"Estoquexxx{data_str}xxxtipo",
		"Exclusividade",
		"codigo","tipo"
	]

def get_campos_alteraveis():
	return json.dumps([
		"Ttl Est","quant_utilizada","quant_pedidos"
	])




def trazer_simulacao(request):

	abre_detalhamento = False if request.COOKIES.get("abre_detalhamento_simulador_de_producao") == "true" else True

	tabela_salva = request.POST.get("simulacoes")
	sqlite_conn = sqlite3.connect("db.sqlite3")

	if not tabela_salva:
		return redirect(reverse("ghost:simulador-de-producao"))

	query = f"SELECT * FROM [simulacao_{tabela_salva}]"

	codigo_aleatorio = gerar_codigo_aleatorio_simulador(10)

	simulador = pd.read_sql(query, sqlite_conn)
	if simulador.empty:
		messages.error(request,"Problemas ao carregar esta tabela")
		return redirect(reverse("ghost:simulador-de-producao"))
	
	simulador.drop(columns=["index"],inplace=True)
	
	_, datas_encontradas = ordenar_colunas_por_data(simulador.filter(like="Produçãoxxx").columns)
	data_str = max(datas_encontradas).strftime("%d-%m-%Y")

	simulador = simulador.fillna('')

	# DATA ESTOQUE
	data_estoque = simulador.filter(like="Estoque").filter(like="codigo").columns[0].split("xxx")[1]

	cabecalhos, rows = get_cabecalhos_e_rows_simulador_de_producao(
		simulador, data_estoque, data_str, data_str,abre_detalhamento
	)
	colunas_fixas = get_colunas_fixas(data_estoque)
	campos_alteraveis = get_campos_alteraveis()

	context = {
		"caller":"trazer_simulacao",
		"cabecalhos":cabecalhos,
		"rows":rows,
		"codigo_aleatorio": codigo_aleatorio,
		"tabela_salva":tabela_salva,
		"colunas_fixas": colunas_fixas,
		"campos_alteraveis": campos_alteraveis,
		"data_str": data_str,
		"data_estoque": data_estoque,
	}

	return render(request, "ghost/simuladordeproducao/simuladordeproducao.html", context)






def get_cabecalhos_e_rows_simulador_de_producao(
		simulador:pd.DataFrame, data_estoque:str, data_str:str, maior_data:str, reduz_campos:bool = True
):
	# CABEÇALHOS
	cat_anterior = ''
	dat_anterior = ''
	cabecalhos = {
		"categoria":[],
		"data":[],
		"campo":[],
	}
	
	if reduz_campos:
		simulador = simulador.loc[
			# (
				(simulador[f"Estoquexxx{data_estoque}xxxtipo"].isin(["PA"]))
			# | (simulador.filter(like=data_str).filter(like="Resultado").iloc[:,0] < 0) |
			# (simulador.filter(like=maior_data).filter(like="Resultado").iloc[:,0] < 0))
			,:
		]
	for col in simulador.columns:
		cat, dat, campo = col.split("xxx")

		if cat != "Produção" or data_str in dat or reduz_campos == False:

			if cat == cat_anterior:
				cabecalhos["categoria"][-1][cat] = (cabecalhos["categoria"][-1][cat][0] + 1, col)
			else:
				cabecalhos["categoria"].append({cat: (1, col)})
			cat_anterior = cat

			if dat == dat_anterior:
				cabecalhos["data"][-1][dat] = (cabecalhos["data"][-1][dat][0] + 1, col)
			else:
				cabecalhos["data"].append({dat: (1, col)})
			dat_anterior = dat

			cabecalhos["campo"].append({campo: (1, col)})

	rows = []
	for i, row in simulador.iterrows():
		if reduz_campos:
			colunas_filtradas = [col for col in simulador.columns if "Produção" not in col or data_str in col]
		else:
			colunas_filtradas = simulador.columns
		row_data = row[colunas_filtradas].to_dict()
		row_data["index"] = i
		rows.append(row_data)


	return cabecalhos, rows




def resultado_pi_pa_produzido(simulador, data_estoque, codigo, data_str, quant):
	if simulador.loc[
		simulador[f"Estoquexxx{data_estoque}xxxcodigo"] == codigo,
		f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
	].empty:
		simulador = pd.concat([
			simulador,
			pd.DataFrame({f"Estoquexxx{data_estoque}xxxcodigo":[codigo]})
		]).sort_values(by=f"Estoquexxx{data_estoque}xxxcodigo",ascending=True)
	simulador.loc[
		simulador[f"Estoquexxx{data_estoque}xxxcodigo"] == codigo,
		f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
	] = quant
	return simulador




def descricao_para_produtos_sem_descricao(simulador,data_estoque,codigo,data_str,quant,engine):
	todos_os_codigos = simulador.loc[
			(simulador[f"Estoquexxx{data_estoque}xxxdescricao"].isna()),
			f"Estoquexxx{data_estoque}xxxcodigo"
		].rename("todos_os_codigos").drop_duplicates()
	if not todos_os_codigos.empty:
		todos_os_codigos = forma_string_codigos(todos_os_codigos)
		info_produtos = get_info_produtos(todos_os_codigos, engine)
		
		simulador = simulador.merge(info_produtos,how="left",left_on=f"Estoquexxx{data_estoque}xxxcodigo",right_on="codigo")

		simulador = simulador.fillna({
			f"Estoquexxx{data_estoque}xxxdescricao": simulador["descricao_produto"],
			f"Estoquexxx{data_estoque}xxxtipo": simulador["tipo_produto"],
			f"Estoquexxx{data_estoque}xxxorigem": simulador["origem_produto"]
		}).drop(columns=["descricao_produto","tipo_produto","origem_produto"])
	info_produtos = None
	return simulador




def verificar_alternativos_dos_itens_negativos(simulador, data_str, data_estoque, codigo, quant):
	
	colunas_resultado_e_estoque = pd.concat([
		simulador.filter(like="Ttl Est"),
		simulador.filter(like="Resultado")
	],axis=1).columns

	col_resultado_anterior = ""
	for col in colunas_resultado_e_estoque:
		if data_str in col:
			break
		col_resultado_anterior = col
	if col_resultado_anterior == "":
		col_resultado_anterior = f"Estoquexxx{data_str}xxxTtl Est"

	negativos_mp = simulador.loc[	
		((simulador[f"Resultadoxxx{data_str}xxxqtd"] < 0) & 
		(~simulador[f"Estoquexxx{data_estoque}xxxtipo"].isin(['PI','PA']))),:
	]
	if not negativos_mp.empty:
		for _, row in negativos_mp.iterrows():
			cod_negativo = row[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"]
			quant_neg = -row[f'Resultadoxxx{data_str}xxxqtd']
			alternativo_de = simulador.loc[
				simulador[f'Produçãoxxx{codigo}_{data_str}_{quant}xxxalternativo_de'] == cod_negativo ,
				:
			]
			if not alternativo_de.empty:
				for _, alt in alternativo_de.sort_values(
					by=f'Produçãoxxx{codigo}_{data_str}_{quant}xxxordem_alt',ascending=True
				).iterrows():
					quant_alt = alt[col_resultado_anterior]
					alternativo = alt[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"]
					if quant_alt > 0:
						if quant_alt >= quant_neg:

							#quant_utilizada
							simulador.loc[
								(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == alternativo),
								f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
							] = simulador.loc[
									(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == alternativo),
									f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
								].fillna(0) - quant_neg
							
							simulador.loc[
								(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
								f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
							] = simulador.loc[
									(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
									f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
								].fillna(0) + quant_neg
							
							# Resultado
							simulador.loc[
								simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == alternativo,
								f"Resultadoxxx{data_str}xxxqtd"
							] = simulador.loc[
									simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == alternativo,
									f"Resultadoxxx{data_str}xxxqtd"
								].fillna(0) - quant_neg
							
							simulador.loc[
								(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
								f"Resultadoxxx{data_str}xxxqtd"
							] = simulador.loc[
									(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
									f"Resultadoxxx{data_str}xxxqtd"
								].fillna(0) + quant_neg
							
							break

						elif quant_alt < quant_neg:

							# quant_utilizada
							simulador.loc[
								(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == alternativo),
								f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
							] = -quant_alt

							simulador.loc[
								(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
								f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
							] = simulador.loc[
									(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
									f"Produçãoxxx{codigo}_{data_str}_{quant}xxxquant_utilizada"
								].fillna(0) + quant_alt
							
							# Resultado
							simulador.loc[
								(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == alternativo),
								f"Resultadoxxx{data_str}xxxqtd"
							] = 0

							simulador.loc[
								(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
								f"Resultadoxxx{data_str}xxxqtd"
							] = simulador.loc[
									(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"] == cod_negativo),
									f"Resultadoxxx{data_str}xxxqtd"
								].fillna(0) + quant_alt
							
							quant_neg -= quant_alt

		negativos_mp = None
		cod_negativo = None
		quant_neg = None
		alternativo_de = None
		quant_alt = None
	return simulador




def phase_out(request):
	context = {
		"caller": "inicial",
	}
	return render(request,"ghost/simuladordeproducao/phase_out.html", context)




def get_cabecalhos_e_rows_phaseout(
		df:pd.DataFrame, cols, reduz_campos:bool = True
):

	cabecalhos = []
	cabecalhos_anteriores = []
	size = len(df.columns[0].split("xxx"))
	for i in range(size):
		cabecalhos.append([])
		cabecalhos_anteriores.append([])

	for col in df.columns:
		for i,subcol in enumerate(col.split("xxx")):
			if subcol == cabecalhos_anteriores[i]:
				cabecalhos[i][-1][subcol] = (cabecalhos[i][-1][subcol][0] + 1, col)
			else:
				cabecalhos[i].append({subcol:(1,col)})
			cabecalhos_anteriores[i] = subcol

	rows = []
	insumo_anterior = ""
	ult_indice = 0
	for i, row in df.iterrows():
		row_data = row.to_dict()
		if row["insumo"] != insumo_anterior:
			rows.append((row_data,1))
			ult_indice = len(rows) - 1
		else:
			rows[ult_indice] = (rows[ult_indice][0],rows[ult_indice][1] + 1)
			for cc in cols:
				row_data.pop(cc)
			rows.append((row_data,1))
		insumo_anterior = row["insumo"]
		


	return cabecalhos, rows




def carregar_estruturas_phase_out(request):

	if request.method != "POST": return redirect(reverse("ghost:phase-out"))

	produtos = request.POST.get("codigos-produtos")

	if not produtos: return redirect(reverse("ghost:phase-out"))

	################ processamento
	codigo_identificador = request.POST.get("codigo-identificador","a")
	processamento = Processamento.objects.filter(codigo_identificador = codigo_identificador)
	if processamento.exists():
		processamento.delete()

	processamento = Processamento.objects.create(
		codigo_identificador = codigo_identificador,
		caller = "phase_out",
		porcentagem = "",
		mensagem1 = "Processando",
		mensagem2 = ""
	)
	################

	engine = get_engine()
	codigo_aleatorio = request.session.get("codigo-aleatorio")

	produtos = produtos.split("\r\n")
	produtos_filtrados = [item for item in produtos if item != ""]

	data_referencia = datetime.today().date()

	estruturas: pd.DataFrame
	request, estruturas, _ = gerar_multiestruturas(
		request= request,
		produtos=produtos_filtrados,
		data_referencia=data_referencia,
		engine=engine,
		caller="phase_out"
	)

	if estruturas.empty:
		messages.info(request,"A consulta retornou sem resultados")
		return redirect(reverse("ghost:phase-out"))
	
	estruturas["ori_alt"] = where(estruturas["alternativo_de"].isna(),"Original","Alternativo")

	estruturas = preencher_info_produto(estruturas, engine)

	estruturas = preencher_qtds_itens_alternativos_phaseout(estruturas, engine)

	estruturas = estruturas.sort_values(by="insumo",ascending=True)
	estruturas = estruturas[["insumo","descricao_insumo","ori_alt","alternativo_de","origem_produto","tipo",
		"codigo_original","descricao_cod_original","quant_utilizada","Exclusividade"]]

	estruturas["Status"] = "Corrente"

	estruturas, colunas_estoque = preencher_estoques_phase_out(estruturas, engine)

	estruturas = preencher_pedidos_phase_out(estruturas, engine)

	estruturas["Saldo Total"] = estruturas["Ttl Est (11, 14, 20)"].fillna(0) + estruturas["Pedidos"].fillna(0)
	estruturas.loc[estruturas["Saldo Total"] == 0,"Saldo Total"] = float("nan")


	# SALVAR NO BD
	request = salvar_dataframe_no_bd(
		request=request,
		df=estruturas,
		inicial_tabela="phaseout",
		codigo_aleatorio=codigo_aleatorio
	)
	codigo_aleatorio = request.session.get("codigo-aleatorio")
	# SALVAR NO BD fim

	estruturas = estruturas.fillna("")

	# cabecalhos, rows = get_cabecalhos_e_rows_dataframe(estruturas)
	colunas_para_mesclar = get_colunas_para_mesclar(estruturas)
	cabecalhos, rows = get_cabecalhos_e_rows_phaseout(estruturas, colunas_para_mesclar)

	context = {
		"caller":"carregar_estruturas",
		"cabecalhos":cabecalhos,
		"rows":rows,
		"colunas_para_mesclar":colunas_para_mesclar,
		"ignorar_js_tabela": True,
		"codigo_aleatorio":codigo_aleatorio,
	}

	return render(request, "ghost/simuladordeproducao/phase_out.html", context)



def preencher_info_produto(df: pd.DataFrame, engine):
	todos_os_produtos = forma_string_codigos(df["insumo"])
	info_produtos = get_info_produtos(todos_os_produtos,engine)
	df = df.merge(
		info_produtos[["codigo","origem_produto","tipo_produto"]],how="left",
		left_on="insumo",
		right_on="codigo"
	).drop(columns=["codigo"]).rename(columns={"tipo_produto":"tipo"})

	return df




def get_colunas_para_mesclar(df: pd.DataFrame):

	colunas_armazem = [x for x in df.columns if len(x) == 2]

	return ["insumo","descricao_insumo","ori_alt","tipo","Exclusividade","origem_produto",*colunas_armazem,"Ttl Est (11, 14, 20)", "custo_medio", "Pedidos", "Saldo Total"]




def carregar_phase_out(request):

	if request.method != "POST": return redirect(reverse("ghost:phase-out"))

	produtos = request.POST.get("codigos-produtos")

	if not produtos: return redirect(reverse("ghost:phase-out"))

	engine = get_engine()
	codigo_aleatorio = request.session.get("codigo-aleatorio")

	if not codigo_aleatorio: return redirect(reverse("ghost:phase-out"))

	produtos = produtos.split("\r\n")
	produtos_filtrados = [str(item).upper().strip() for item in produtos if item != ""]

	data_referencia = datetime.today().date()

	# PEGAR A TABELA SALVA NO BANCO
	sqlite_conn = sqlite3.connect('db.sqlite3')
	estruturas = pd.read_sql(f"SELECT * FROM [{codigo_aleatorio}]",sqlite_conn).drop(columns="index")
	# if not estruturas.empty:
	# 	cursor = sqlite_conn.cursor()
	# 	cursor.execute(f"DROP TABLE [{codigo_aleatorio}]")
	# 	sqlite_conn.commit()
	sqlite_conn.close()
	
	# DUPLICIDADE PRODUTO
	estruturas = estruturas.loc[~estruturas["codigo_original"].isin(produtos_filtrados),:]


	phouts: pd.DataFrame
	request, phouts, _ = gerar_multiestruturas(
		request= request,
		produtos=produtos_filtrados,
		data_referencia=data_referencia,
		engine=engine,
		caller="phase_out"
	)

	phouts["ori_alt"] = where(phouts["alternativo_de"].isna(),"Original","Alternativo")

	phouts = preencher_info_produto(phouts, engine)

	phouts = preencher_qtds_itens_alternativos_phaseout(phouts, engine)

	phouts = phouts[["insumo","descricao_insumo","ori_alt","alternativo_de","origem_produto","tipo",
		"codigo_original","descricao_cod_original","quant_utilizada","Exclusividade"]]

	# EXCLUSIVIDADE PHASE OUTS
	phouts_comuns = phouts.loc[phouts["insumo"].isin(estruturas["insumo"]),["insumo"]]
	if not phouts_comuns.empty:
		phouts.loc[phouts["insumo"].isin(phouts_comuns["insumo"]),"Exclusividade"] = "COMUM"
		estruturas.loc[estruturas["insumo"].isin(phouts_comuns["insumo"]),"Exclusividade"] = "COMUM"
	phouts_comuns = None

	phouts.loc[
		((~phouts["insumo"].isin(estruturas["insumo"])) & 
		(phouts["Exclusividade"] == "COMUM")),"Exclusividade"
	] = "COMUM ENTRE PHASE OUTS"

	phouts["Status"] = "Phase Out"

	phouts, colunas_estoque = preencher_estoques_phase_out(phouts,engine)

	phouts = preencher_pedidos_phase_out(phouts, engine)

	estru_phouts = pd.concat([estruturas, phouts])

	colunas_armazem = sorted([x for x in estru_phouts.columns if len(x) == 2])
	estru_phouts = estru_phouts[[
		"insumo","descricao_insumo","ori_alt","alternativo_de","codigo_original","descricao_cod_original",
		"quant_utilizada","Exclusividade","Status","origem_produto","tipo",*colunas_armazem,"Ttl Est (11, 14, 20)",
		"custo_medio","Pedidos"
	]]

	estru_phouts["Saldo Total"] = estru_phouts["Ttl Est (11, 14, 20)"].fillna(0) + estru_phouts["Pedidos"].fillna(0)
	estru_phouts.loc[estru_phouts["Saldo Total"] == 0,"Saldo Total"] = float("nan")

	estru_phouts = estru_phouts.sort_values(by=["insumo","codigo_original"],ascending=[True,True])

	request = salvar_dataframe_no_bd(
		request=request,
		df=estru_phouts,
		codigo_aleatorio=codigo_aleatorio
	)

	estru_phouts = estru_phouts.fillna("")

	colunas_para_mesclar = get_colunas_para_mesclar(estru_phouts)
	cabecalhos, rows = get_cabecalhos_e_rows_phaseout(estru_phouts, colunas_para_mesclar)

	# gerar_simulacao_excel(codigo_aleatorio,cabecalhos,rows,colunas_para_mesclar, colunas_estoque)
	# relatorio_phaseout_com_openpyxl(codigo_aleatorio,cabecalhos,rows,colunas_para_mesclar)
	# relatorio_phaseout_por_produto(codigo_aleatorio,colunas_para_mesclar)

	context = {
		"caller":"carregar_phase_out",
		"cabecalhos":cabecalhos,
		"rows":rows,
		"colunas_para_mesclar":colunas_para_mesclar,
		"ignorar_js_tabela": True,
		"codigo_aleatorio": codigo_aleatorio,
	}

	return render(request, "ghost/simuladordeproducao/phase_out.html", context)




def preencher_produtos_sem_estoque(
		simulador, codigo, data_str, quant, data_estoque
):
	produtos_sem_estoque = simulador.loc[simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"].isna(),:]
	if not produtos_sem_estoque.empty:
		# simulador.replace({f"Estoquexxx{hoje}xxxcodigo":{"":nan}},inplace=True)
		simulador.loc[:,f"Estoquexxx{data_estoque}xxxcodigo"] = simulador[f"Estoquexxx{data_estoque}xxxcodigo"]\
			.combine_first(simulador[f"Produçãoxxx{codigo}_{data_str}_{quant}xxxinsumo"])

	return simulador




def padronizar_cabecalhos_estoque(estoque_pivot, data_estoque):
	novos_cabecalhos = {}
	for col in estoque_pivot.columns:
		novos_cabecalhos.update({col:f"Estoquexxx{data_estoque}xxx{col}"})

	estoque_pivot = estoque_pivot.rename(columns=novos_cabecalhos)
	return estoque_pivot




def padronizar_cabecalhos_pedidos(pedidos_pivot):
	novos_cabecalhos = {}
	for col in pedidos_pivot.columns:
		novos_cabecalhos.update({col:f"Pedidosxxx{col}xxxquant_pedidos"})

	pedidos_pivot = pedidos_pivot.rename(columns=novos_cabecalhos)
	return pedidos_pivot




def get_pedidos_pivot(engine,coluna_codigos,abre_datas = True):
	pedidos = get_pedidos(
		solicitante="simulador",
		engine=engine
	)
	pedidos_filtrados = pedidos.loc[
		pedidos[f"codigo"].isin(coluna_codigos.drop_duplicates()),:
	]
	if abre_datas:
		pedidos_pivot = pedidos_filtrados.pivot_table(
			index="codigo", 
			columns="entrega", 
			values="quant",
			aggfunc="sum"
		).reset_index()
	else:
		pedidos_pivot = pedidos_filtrados.pivot_table(
			index="codigo",
			values="quant",
			aggfunc="sum"
		).reset_index()
	return pedidos_pivot




def salvar_dataframe_no_bd(request,df,inicial_tabela = "",codigo_aleatorio=None):
	if not codigo_aleatorio:
		codigo_aleatorio = gerar_codigo_aleatorio_simulador(10,inicial_tabela)
		request.session["codigo-aleatorio"] = codigo_aleatorio
	print(codigo_aleatorio)
	sqlite_conn = sqlite3.connect('db.sqlite3')
	df.to_sql(name=codigo_aleatorio, con=sqlite_conn, if_exists="replace", index=True)
	sqlite_conn.close()
	return request




def preencher_qtds_itens_alternativos_phaseout(df:pd.DataFrame, engine):
	df["quant_utilizada"] = df["quant_utilizada"].fillna(
		df.merge(df,how="left",
		   left_on=["alternativo_de","codigo_original"],
		   right_on=["insumo","codigo_original"],
		   suffixes=("","_alt")
		)["quant_utilizada_alt"]
	)

	codigos_sem_descricao = df.loc[df["descricao_insumo"].isna(),"insumo"].drop_duplicates()
	codigos_str = forma_string_codigos(codigos_sem_descricao)

	info_produtos = get_info_produtos(codigos=codigos_str, engine=engine)

	df["descricao_insumo"] = df["descricao_insumo"].fillna(
		df.merge(
			info_produtos,how="left",
			left_on="insumo",
			right_on="codigo"
		)["descricao_produto"]
	)

	return df




def preencher_estoques_phase_out(df:pd.DataFrame, engine):
		# ESTOQUE
	todos_os_codigos = forma_string_codigos(df["insumo"].drop_duplicates())

	query_estoque = get_query_estoque_atual()
	estoque = pd.read_sql(text(query_estoque),engine, params={
		"codigos": todos_os_codigos,
	})

	# CALCULA O CUSTO MÉDIO PONDERADO
	estoque["total"] = estoque["quant"].fillna(0) * estoque["unitario"].fillna(0)
	custo_medio = estoque.groupby(by="codigo").apply(lambda x: round( x["total"].fillna(0).sum() / x["quant"].fillna(0).sum() ,5)).reset_index(name='custo_medio')

	# FORMA O ESTOQUE_PIVOT
	estoque_pivot = estoque.pivot(index=["codigo","descricao"], columns="armazem", values="quant").reset_index()
	arm_exist = ["11","14","20"]
	arm_verificados = [arm for arm in arm_exist if arm in estoque_pivot.columns]
	estoque_pivot["Ttl Est (11, 14, 20)"] = estoque_pivot[arm_verificados].sum(axis=1).round(5)

	estoque_pivot = estoque_pivot.merge(custo_medio, on="codigo")

	df = df.merge(
		estoque_pivot,how="left",
		left_on="insumo",
		right_on="codigo"
	).drop(columns=["codigo","descricao"])

	colunas = list(estoque_pivot.drop(columns=["codigo","descricao"]).columns)

	return df, colunas




def preencher_pedidos_phase_out(df:pd.DataFrame, engine):
	# PEDIDOS
	pedidos_pivot = get_pedidos_pivot(
		engine=engine,
		coluna_codigos=df["insumo"],
		abre_datas=False
	)

	df = df.merge(
		pedidos_pivot,how="left",
		left_on="insumo",
		right_on="codigo"
	).drop(columns=["codigo"]).rename(columns={"quant":"Pedidos"})

	return df




def gerar_simulacao_excel(codigo_aleatorio,cabecalhos,rows,colunas_para_mesclar, colunas_estoque):
	app = xw.App(visible=False)
	wb = app.books.add()

	try:

		ws = wb.sheets[0]

		prim_lin = 2
		prim_col = 2

		ws.range((1,prim_col)).value = "Simulação Phase Out"
		ws.cells(1,prim_col).api.Font.Bold = True

		lin = prim_lin
		full_cabs = []
		for nivel_cab in cabecalhos:
			col = prim_col
			for cabecalho in nivel_cab:
				for cab, info_cab in cabecalho.items():
					colspan, full_cab = info_cab
					celula = ws.cells(lin,col)
					celula.value = cab
					celula.api.Font.Bold = True
					celula.color = (50,50,50)
					celula.api.Font.Color = rgb_para_long((255,255,255))
					# celula.api.Borders.LineStyle = 1
					if colspan > 1:
						ws.range((lin,col),(lin,col+colspan-1)).merge()
						col = col+colspan-1
					if lin == prim_lin:
						full_cabs.append(full_cab)
					col += 1
			lin += 1

		for row, rowspan in rows:
			col = prim_col
			for cab in full_cabs:
				if row.get(cab):
					ws.cells(lin,col).value = row[cab]
					# ws.cells(lin,col).api.Borders.LineStyle = 1
				if cab in colunas_para_mesclar and rowspan > 1:
					ws.range((lin,col),(lin+rowspan-1,col)).merge()
					ws.cells(lin,col).api.VerticalAlignment = -4108
				col += 1
			lin += 1

		ws.range(f"{gcl(prim_col)}:{gcl(col-1)}").autofit()
		ws.range((prim_lin,prim_col),(lin-1,col-1)).api.Borders.LineStyle = 1

		# range_tabela = ws.range((prim_lin,prim_col),(lin-1,col-1))
		# tabela = ws.api.ListObjects.Add(1,range_tabela.api, "", 1)
		# tabela.Name = "simulador"



		caminho = path.join(settings.MEDIA_ROOT,f"{codigo_aleatorio}.xlsx")
		wb.save(caminho)
	except Exception as e: 
		print(e)
	finally:
		wb.close()
		app.quit()




def relatorio_phaseout_com_openpyxl(codigo_aleatorio,cabecalhos,rows,colunas_para_mesclar):
	wb = Workbook()
	ws = wb.active
	ws.title = "Simulador"

	prim_lin = 2
	prim_col = 2

	lin = prim_lin
	full_cabs = []
	for nivel_cab in cabecalhos:
		col = prim_col
		for cabecalho in nivel_cab:
			for cab, info_cab in cabecalho.items():
				colspan, full_cab = info_cab
				celula = ws.cell(lin,col,cab)
				if colspan > 1:
					celula.comment = Comment(f"lin:{colspan}","")
				if lin == prim_lin:
					full_cabs.append(full_cab)
				col += 1
		lin += 1

	for row, rowspan in rows:
		col = prim_col
		for cab in full_cabs:
			if row.get(cab):
				celula = ws.cell(lin, col, row[cab])
			if cab in colunas_para_mesclar and rowspan > 1:
				celula.comment = Comment(f"col:{rowspan}","")
			col += 1
		lin += 1
	
	print("xlwings")

	caminho = path.join(settings.MEDIA_ROOT,f"{codigo_aleatorio}_.xlsx")
	wb.save(caminho)

	app = xw.App(visible=False,add_book=False)
	xwwb = app.books.open(caminho)
	xwws = xwwb.sheets("Simulador")

	for l in range(prim_lin,lin,1):
		for c in range(prim_col,col,1):
			if xwws.cells(l,c).api.Comment:
				direcao, comm = xwws.cells(l,c).api.Comment.Text().split(":")
				comm = int(comm)
				if direcao == "lin":
					xwws.range((l,c),(l,c+comm-1)).merge()
				elif direcao == "col":
					xwws.range((l,c),(l+comm-1,c)).merge()
				xwws.cells(l,c).api.VerticalAlignment = -4108
				xwws.cells(l,c).api.Comment.Delete()
	
	xwwb.save()
	xwwb.close()
	app.quit()




@api_view(["POST"])
def relatorio_phaseout_por_produto(request):

	data = request.data
	codigo_identificador = data.get("codigo_processamento")
	mesclar = True
	porcent = 20 if mesclar else 79

	################ processamento
	processamento = Processamento.objects.filter(codigo_identificador=codigo_identificador)
	if processamento.exists():
		processamento.delete()

	processamento = Processamento.objects.create(
		codigo_identificador = data.get("codigo_processamento","a"),
		caller = "phase_out",
		porcentagem = "",
		mensagem1 = "Processando",
		mensagem2 = "Preparando informações"
	)
	################

	codigo_aleatorio = data.get("codigo_aleatorio")

	qtd_meses = 6
	df:pd.DataFrame
	sqlite_conn = sqlite3.connect("db.sqlite3")
	df = pd.read_sql(f"SELECT * FROM [{codigo_aleatorio}]",sqlite_conn).drop(columns="index")
	sqlite_conn.close()

	if df.empty:
		return Response({"sucesso":False,"erro":"Consulta Vazia"})

	colunas_para_mesclar = get_colunas_para_mesclar(df)

	wb = Workbook()
	ws = wb.active
	ws2 = wb.create_sheet(title="Demanda")
	ws.title = "Simulador"

	produtos = df.loc[:,"insumo"].drop_duplicates()
	originais = df.loc[:,"codigo_original"].drop_duplicates()
	prim_lin = 2
	prim_col = 2

	# CORES
	cinza_claro = Color(rgb_para_hex(235,235,235))
	cinza_escuro = Color(rgb_para_hex(50,50,50))
	cinza_menos_escuro = Color(rgb_para_hex(140,140,140))
	branco = Color(rgb_para_hex(255,255,255))
	vermelho_escuro_cor = Color(rgb_para_hex(156,0,6))
	vermelho_claro_cor = Color(rgb_para_hex(255,199,206))
	verde_claro_cor = Color(rgb_para_hex(198,239,206))
	verde_escuro_cor = Color(rgb_para_hex(0,97,0))
	amarelo_escuro_cor = Color(rgb_para_hex(156,87,0))
	amarelo_claro_cor = Color(rgb_para_hex(255,235,156))

	# BORDAS
	borda_fina_style = Side(style="dashDot")
	borda_pontilhada_style = Side(style="dashDot")
	borda_pont_branca_style = Side(style="dashDot", color=f"{rgb_para_hex(255,255,255)}")
	borda_fina = Border(
		left=borda_fina_style,
		right=borda_fina_style,
		top=borda_fina_style,
		bottom=borda_fina_style
	)
	borda_grossa_inferior = Border(
		left=borda_fina_style,
		right=borda_fina_style,
		top=borda_fina_style,
		bottom=Side(style="thick")
	)
	borda_pontilhada_preta = Border(
		left=borda_pontilhada_style,
		right=borda_pontilhada_style,
		top=borda_pontilhada_style,
		bottom=borda_pontilhada_style
	)
	borda_pontilhada_branca = Border(
		left=borda_pont_branca_style,
		right=borda_pont_branca_style,
		top=borda_pont_branca_style,
		bottom=borda_pont_branca_style
	)

	# FUNDOS
	fundo_escuro = PatternFill(start_color=cinza_escuro,patternType="solid")
	fundo_menos_escuro = PatternFill(start_color=cinza_menos_escuro,patternType="solid")
	fundo_vermelho_claro = PatternFill(start_color=vermelho_claro_cor,end_color=vermelho_claro_cor,patternType="solid")
	fundo_verde_claro = PatternFill(start_color=verde_claro_cor,end_color=verde_claro_cor,patternType="solid")
	fundo_amarelo_claro = PatternFill(start_color=amarelo_claro_cor,end_color=amarelo_claro_cor,patternType="solid")


	# FONTES
	fonte_branca = Font(color=branco,b=True)
	fonte_verde = Font(color=verde_escuro_cor)
	fonte_vermelha = Font(color=vermelho_escuro_cor)
	fonte_amarela = Font(color=amarelo_escuro_cor)

	# REGRAS
	regra_vermelho = CellIsRule(operator="lessThan",formula=["0"],font=fonte_vermelha,fill=fundo_vermelho_claro)
	regra_verde = CellIsRule(operator="greaterThan",formula=["0"],font=fonte_verde,fill=fundo_verde_claro)
	regra_amarelo = CellIsRule(operator="equal",formula=["0"],font=fonte_amarela,fill=fundo_amarelo_claro)


	lin = prim_lin
	col = prim_col
	cab_col = {}
	mesclagem = []
	cols_para_formatacao_condicional = set()

	################ processamento
	processamento.mensagem2 = "Imputando Cabeçalhos"
	processamento.save()
	################

	def estilo_celula_cabecalho(celula):
		celula.fill = fundo_escuro
		celula.font = fonte_branca
		celula.border = borda_pontilhada_branca

	for cabecalho in df.columns:
		celula = ws.cell(lin,col,cabecalho)
		estilo_celula_cabecalho(celula)
		cab_col.update({cabecalho:col})
		if cabecalho == "custo_medio":
			coluna_custo_medio = col # para a fórmula da ultima coluna
		col += 1

	prim_lin_ws2 = lin_ws2 = 2
	prim_col_ws2 = col_ws2 = 2
	ws2.cell(lin_ws2,col_ws2,"Produto")
	for ori in originais:
		lin_ws2 += 1
		ws2.cell(lin_ws2,col_ws2,ori)
	ult_lin_ws2 = lin_ws2
	col_ws2 += 1

		# CABEÇALHOS DOS MESES
	locale.setlocale(locale.LC_TIME, "Portuguese_Brazil.1252")
	mes_atual = datetime.now()
	for m in range(qtd_meses):

		mes_atual_str = mes_atual.strftime("%B/%y")
		ws2.cell(prim_lin_ws2,col_ws2,mes_atual_str)
		for lin_ws2 in range(prim_lin_ws2+1,ult_lin_ws2+1):
			ws2.cell(lin_ws2,col_ws2,0)

		celula = ws.cell(lin-1,col,mes_atual_str)
		estilo_celula_cabecalho(celula)
		mesclagem.append(f"{gcl(col)}{lin-1}:{gcl(col+2)}{lin-1}")

		celula = ws.cell(lin,col,"Saldo Inicial")
		estilo_celula_cabecalho(celula)
		cols_para_formatacao_condicional.add(col)
		col += 1

		celula = ws.cell(lin, col, "Necessidade")
		estilo_celula_cabecalho(celula)
		col += 1

		celula = ws.cell(lin, col, "Saldo Final")
		estilo_celula_cabecalho(celula)
		cols_para_formatacao_condicional.add(col)
		col += 1

		mes_atual = mes_atual + relativedelta(months=1)
		col_ws2 += 1
	
	celula = ws.cell(lin,col,"E&O (R$)")
	estilo_celula_cabecalho(celula)

	def estilo_celula_corpo(celula, cor_da_vez = None):
		if cor_da_vez:
			celula.fill = PatternFill(start_color=cor_da_vez,patternType="solid")
		celula.border = borda_fina
	
	def adicionar_formatacao_condicional(coord):
		ws.conditional_formatting.add(coord,regra_amarelo)
		ws.conditional_formatting.add(coord,regra_verde)
		ws.conditional_formatting.add(coord,regra_vermelho)

	lin += 1

	df_prod: pd.DataFrame
	cor_da_vez = cinza_claro
	index = 0
	for produto in produtos:
		
		index += 1

		################ processamento
		processamento.porcentagem = f'{int((index+1)/len(produtos)*porcent)}%'
		processamento.mensagem2 = f'Imprimindo {produto}'
		processamento.save()
		################
		
		df_prod  = df.loc[df["insumo"] == produto,:]
		qtd_linhas = df_prod.shape[0]

		if cor_da_vez == cinza_claro:
			cor_da_vez = branco
		else:
			cor_da_vez = cinza_claro
		
		linini = lin
		for _, row in df_prod.iterrows():
			col = prim_col
			for cab in df.columns:
				celula = ws.cell(lin,col,row[cab])
				estilo_celula_corpo(celula, cor_da_vez)
				col += 1
			
			prim_col_formula = col
			
			for m in range(qtd_meses):
				if lin == linini:
					if col == prim_col_formula:
						formula = f"={gcl(col-1)}{lin}"
						celula = ws.cell(lin,col,formula)
					else:
						formula = f"={gcl(col-1)}{lin+qtd_linhas}"
						celula = ws.cell(lin,col,formula)
				else:
						formula = f"={gcl(col+2)}{lin-1}"
						celula = ws.cell(lin,col,formula)
				estilo_celula_corpo(celula)
				col += 1
			
				formula = f"=IF({gcl(cab_col['ori_alt'])}{linini}=\"Alternativo\",0,"\
					f"VLOOKUP({gcl(cab_col['codigo_original'])}{lin},'{ws2.title}'!{gcl(prim_col_ws2)}:"\
					f"{gcl(m+prim_col_ws2+1)},{m+2},0)*{gcl(cab_col['quant_utilizada'])}{lin})"
				celula = ws.cell(lin,col,formula)
				estilo_celula_corpo(celula)
				col += 1

				formula = f"={gcl(col-2)}{lin}-{gcl(col-1)}{lin}"
				celula = ws.cell(lin,col,formula)
				estilo_celula_corpo(celula)

				col += 1

			lin += 1

		col = prim_col_formula

		for c in range(prim_col, prim_col_formula):
			ws.cell(lin,c).fill = fundo_menos_escuro

		for m in range(qtd_meses):
			formula = f"={gcl(col-1)}{linini}"
			celula = ws.cell(lin,col,formula)
			col += 1

			formula = f"=SUM({gcl(col)}{linini}:{gcl(col)}{lin-1})"
			celula = ws.cell(lin,col,formula)
			celula.fill = PatternFill(start_color=cinza_claro,patternType="solid")
			col += 1

			formula = f"={gcl(col-2)}{lin}-{gcl(col-1)}{lin}"
			celula = ws.cell(lin,col,formula)
			col += 1
		
		formula = f"=IF({gcl(coluna_custo_medio)}{lin-1}*{gcl(col-1)}{lin}<0,0,{gcl(coluna_custo_medio)}{lin-1}*{gcl(col-1)}{lin})"
		celula = ws.cell(lin,col,formula)
		adicionar_formatacao_condicional(celula.coordinate)
		
		
		#mesclagem
		if qtd_linhas > 1:
			for col_mescla in colunas_para_mesclar:
				letra_col_mescla = gcl(cab_col[col_mescla])
				if col_mescla in ["insumo","descricao_insumo"]:
					mesclagem.append(f"{letra_col_mescla}{linini}:{letra_col_mescla}{lin}")
				else:
					mesclagem.append(f"{letra_col_mescla}{linini}:{letra_col_mescla}{lin-1}")
		else:
			mesclagem.append(f"{gcl(cab_col['insumo'])}{lin-1}:{gcl(cab_col['insumo'])}{lin}")
			mesclagem.append(f"{gcl(cab_col['descricao_insumo'])}{lin-1}:{gcl(cab_col['descricao_insumo'])}{lin}")

		for celula in ws[f"{gcl(prim_col)}{lin}:{gcl(col)}{lin}"][0]:
			celula.border = borda_grossa_inferior
		
		lin += 1
	
	for c in cols_para_formatacao_condicional:
		cc = gcl(c)
		str_formatacao_condicional = f"{cc}{prim_lin+1}:{cc}{lin-1}"
		adicionar_formatacao_condicional(str_formatacao_condicional)
	

	
	caminho = path.join(settings.MEDIA_ROOT,f"{codigo_aleatorio}.xlsx")
	wb.save(caminho)

	if mesclar:

		app = xw.App(visible=False,add_book=False)
		xwwb = app.books.open(caminho)
		xwws = xwwb.sheets("Simulador")

		################ processamento
		processamento.mensagem2 = f'Mesclando células'
		processamento.save()
		################
		
		index = 0
		for mescla in mesclagem:

			index += 1
			################ processamento
			processamento.porcentagem = f'{int((index+1)/len(mesclagem)*porcent)+20}%'
			processamento.save()
			################
			
			xwws.range(mescla).merge()
			xwws.range(mescla).api.VerticalAlignment = -4108
			xwws.range(mescla).api.HorizontalAlignment = constants.xlCenter
		
		xwws.range(f"{gcl(prim_col)}:{gcl(col - 1)}").autofit()

		################ processamento
		processamento.porcentagem = '100%'
		processamento.mensagem1 = 'Concluído'
		processamento.mensagem2 = ''
		processamento.save()
		################
		
		xwwb.save()
		xwwb.close()
		app.quit()

	with open(caminho, "rb") as file:

		response = HttpResponse(
			file.read(),
			content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
		)
		response["Content-Disposition"] = f'attachment; filename="{codigo_aleatorio}.xlsx"'
	return response