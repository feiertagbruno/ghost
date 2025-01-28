from dotenv import load_dotenv
from os import environ, startfile
from sqlalchemy import create_engine, text
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from ghost.queries import *



def tratamento_data_referencia(data_referencia):
	if not data_referencia:
		data_referencia = datetime.today().date()
	elif type(data_referencia) == str:
		if data_referencia.count("-") == 2:
			data_referencia = datetime.strptime(data_referencia, "%Y-%m-%d").date()
		elif data_referencia.count("/") == 2:
			data_referencia = datetime.strptime(data_referencia, "%d/%m/%Y").date()
		else:
			data_referencia = datetime.strptime(data_referencia, "%Y%m%d").date()
	return data_referencia




def get_engine():
	load_dotenv()

	SERVER = environ.get("SERVER")
	DB = environ.get("DB")
	USER = environ.get("USER")
	PWD = environ.get("PWD")
	DRIVER = "ODBC Driver 17 for SQL Server"

	connection_string = f"mssql+pyodbc://{USER}:{PWD}@{SERVER}/{DB}?driver={DRIVER}"

	engine = create_engine(connection_string)
	
	return engine




def get_estrutura_produto(codigo, data_referencia, engine):

	query = get_query_estrutura_produto()
	resultado = pd.read_sql(text(query), engine, params={
		"data_referencia": data_referencia,
		"codigo": codigo,
	})

	return resultado




def forma_string_codigos(codigos):
	str_codigos = ""
	for cod in codigos:
		str_codigos += f"{cod},"
	return str_codigos.rstrip(",")





def acrescenta_alternativos(estrutura: pd.DataFrame, engine):
	codigos_filhos = estrutura[estrutura["tipo_insumo"] != "PI"][["insumo"]].reset_index(drop=True)
	todos_os_codigos = codigos_filhos.copy(deep=True).rename(columns={"insumo":"todos_os_codigos"})

	str_codigos = forma_string_codigos(codigos_filhos["insumo"])

	query = get_query_alternativos()

	resultado = pd.read_sql(text(query), engine, params={
		"codigos": str_codigos,
	})

	todos_os_codigos = pd.concat([todos_os_codigos,(
		resultado.drop_duplicates(subset="alternativos")[["alternativos"]].rename(columns={"alternativos":"todos_os_codigos"})
	)],ignore_index=True)
	todos_os_codigos.drop_duplicates(subset="todos_os_codigos",inplace=True)

	resultado = resultado.groupby("prodori")["alternativos"].agg(lambda x: ";".join(x)).reset_index()
	estrutura = estrutura.merge(resultado[["prodori","alternativos"]],how="left",left_on="insumo",right_on="prodori")\
		.drop(columns="prodori").fillna(value="")

	return estrutura, todos_os_codigos




def busca_custos_ultima_compra(str_codigos, data_referencia, engine):

	query = get_query_ultima_compra_produtos()

	return pd.read_sql(text(query), engine, params={
		"codigos":str_codigos,
		"data_referencia": data_referencia,
	})

	#começar pelo preço da última compra




def explode_extrutura(codigo, data_referencia = None, engine = None):

	data_referencia = tratamento_data_referencia(data_referencia)
	
	if not engine:
		engine = get_engine()

	#primeira consulta no banco
	resultado = get_estrutura_produto(codigo, data_referencia, engine)
	estrutura = resultado.copy(deep=True)

	filtro_PI = resultado[resultado["tipo_insumo"] == "PI"]
	tem_PI = True if not filtro_PI.empty else False

	#enquanto tiver PI continua executando a consulta no banco
	while tem_PI:
		
		for codigo_PI in filtro_PI["insumo"]:
			resultado = get_estrutura_produto(codigo_PI, data_referencia, engine)
			estrutura.loc[estrutura["insumo"] == codigo_PI, "verificado"] = True
			estrutura = pd.concat([estrutura,resultado],ignore_index=True)

		filtro_PI = estrutura[(estrutura["verificado"].isnull()) & (estrutura["tipo_insumo"] == "PI")]
		tem_PI = True if not filtro_PI.empty else False

	resultado = None
	filtro_PI = None
	tem_PI = None

	estrutura["quant_utilizada"] = estrutura.groupby("insumo")["quant_utilizada"].transform("sum")
	estrutura.drop_duplicates(subset="insumo", inplace=True)
	
	estrutura, todos_os_codigos = acrescenta_alternativos(estrutura, engine)

	return estrutura, todos_os_codigos

	# FIM explode_extrutura




def busca_custo_e_comentario_alternativos(row, consulta_custos, coluna_custo, coluna_custo_alt, coluna_comentario, coluna_comentario_alt):
	alternativos = row["alternativos"].split(";")
	
	if not alternativos[0]: 
		return pd.Series({
		coluna_custo_alt: "",
		coluna_comentario_alt: ""
	})

	custos_alternativos = [
		str(consulta_custos.loc[consulta_custos["insumo"] == alt, coluna_custo].values[0] * row["quant_utilizada"] ).replace(".",",")
		for alt in alternativos if alt in consulta_custos["insumo"].values
	]
	comentarios_alternativos = [
		consulta_custos.loc[consulta_custos["insumo"] == alt, coluna_comentario].values[0]
		for alt in alternativos if alt in consulta_custos["insumo"].values
	]

	return pd.Series({
		coluna_custo_alt: ";".join(custos_alternativos) if custos_alternativos != [None] else "",
		coluna_comentario_alt: "\n\n".join(comentarios_alternativos) if comentarios_alternativos != [None] else ""
	})




def traz_custos_por_produto(estrutura, consulta_custos, tupla_nomes_colunas):
	coluna_custo, coluna_comentario = tupla_nomes_colunas

	estrutura = estrutura.merge(consulta_custos[["insumo",coluna_custo,coluna_comentario]], on="insumo", how="left")\
		.assign(**{coluna_custo: lambda x: x["quant_utilizada"] * x[coluna_custo]}).fillna(0)

	if "alternativos" in estrutura.columns:

		coluna_custo_alt = f"{coluna_custo}_alt"
		coluna_comentario_alt = f"{coluna_comentario}_alt"
		estrutura[coluna_custo_alt] = ""
		estrutura[coluna_comentario_alt] = ""

		estrutura[[coluna_custo_alt,coluna_comentario_alt]] = estrutura.apply(
			busca_custo_e_comentario_alternativos, 
			axis=1, 
			args=(consulta_custos, coluna_custo, coluna_custo_alt, coluna_comentario, coluna_comentario_alt)
		)
	
	return estrutura




def calcula_custo_total(codigo, descricao, data_referencia, estrutura:pd.DataFrame, nomes_colunas, custos_totais_produto = pd.DataFrame()):
	coluna_custos_estrutura, coluna_custos_alternativos, nova_coluna_custo_total, nova_coluna_comentario = nomes_colunas
	
	if custos_totais_produto.empty:
		custos_totais_produto = pd.DataFrame([{
			"codigo_original": codigo,
			"descricao": descricao,
			"data_referencia": data_referencia,
			nova_coluna_custo_total: None,
			nova_coluna_comentario: None,
		}])
	elif custos_totais_produto[custos_totais_produto["codigo_original"] == codigo].empty:
		custos_totais_produto.loc[len(custos_totais_produto)] = [
			codigo, estrutura.loc[estrutura["codigo_pai"] == codigo, "descricao_pai"].values[0], data_referencia
		]

	comentario = f"{codigo}\nData Referência: {datetime.strftime(data_referencia,'%d/%m/%Y')}\n\n"
	#calculo dos originais (os únicos no caso de OPs)
	estrutura_codigos_originais = estrutura[
		(estrutura[coluna_custos_estrutura].notnull()) & (estrutura[coluna_custos_estrutura] != 0)
	][["insumo","descricao_insumo",coluna_custos_estrutura]]

	total = estrutura_codigos_originais[coluna_custos_estrutura].sum()
	comentario += estrutura_codigos_originais[["insumo","descricao_insumo",coluna_custos_estrutura]].agg(
		lambda x: f"Ori - {x['insumo']} - {x['descricao_insumo']}: R$ {str(round(x[coluna_custos_estrutura],5)).replace('.',',')}"
		,axis=1
	).str.cat(sep="\n")

	produtos_nao_encontrados = estrutura[
		(estrutura[coluna_custos_estrutura].isnull()) | (estrutura[coluna_custos_estrutura] == 0)
	][["insumo","descricao_insumo"]]


	#quando tem a coluna alternativos
	if "alternativos" in estrutura.columns:
		estrutura_alternativos = estrutura[(
			((estrutura[coluna_custos_estrutura].isnull()) | (estrutura[coluna_custos_estrutura] == 0)) &
			((estrutura[coluna_custos_alternativos].notnull()) & (estrutura[coluna_custos_alternativos] != ""))
		)][["alternativos", coluna_custos_alternativos]]
		produtos_nao_encontrados = estrutura[(
			((estrutura[coluna_custos_estrutura].isnull()) | (estrutura[coluna_custos_estrutura] == 0)) &
			((estrutura[coluna_custos_alternativos].isnull()) | (estrutura[coluna_custos_alternativos] == "")) & 
			((estrutura["alternativos"].notnull()) & estrutura["alternativos"] != "")
		)][["insumo","descricao_insumo"]]

		if not estrutura_alternativos.empty:
			total += estrutura_alternativos[coluna_custos_alternativos].str.split(";").str[0].str.replace(",",".").astype(float).sum()
			comentario += estrutura_alternativos[["alternativos",coluna_custos_alternativos]].agg(
				lambda x: f"Alt-{x['alternativos'].split(';')[0]}: R$ {x[coluna_custos_alternativos].split(';')[0]}"
				,axis=1
			).str.cat(sep="\n")

	#produtos não encontrados
	comentario += "\n\nProdutos não encontrados:\n"
	produtos_nao_encontrados[["insumo", "descricao_insumo"]] = produtos_nao_encontrados[["insumo", "descricao_insumo"]].astype(str)
	comentario += (produtos_nao_encontrados["insumo"] + " - " + produtos_nao_encontrados["descricao_insumo"]).str.cat(sep="\n")
	
	custos_totais_produto.loc[custos_totais_produto["codigo_original"] == codigo, nova_coluna_custo_total] = round(total,5)
	custos_totais_produto.loc[custos_totais_produto["codigo_original"] == codigo, nova_coluna_comentario] = comentario

	return custos_totais_produto




def busca_custos_ultimo_fechamento(str_codigos, data_referencia, engine):
	query = get_query_ultimo_fechamento_produtos()
	
	return pd.read_sql(text(query), engine, params={
		"codigos": str_codigos,
		"data_referencia": data_referencia,
	})




def busca_custos_medios(str_codigos, data_referencia, engine):
	query = get_query_custos_medios_produtos()
	
	return pd.read_sql(text(query), engine, params={"codigos":str_codigos})



def estrutura_simples(codigo, data_referencia):

	engine = get_engine()
	data_referencia = tratamento_data_referencia(data_referencia)

	estrutura, todos_os_codigos = explode_extrutura(codigo, data_referencia, engine)
	estrutura = estrutura[estrutura["tipo_insumo"] != "PI"]

	str_codigos = forma_string_codigos(todos_os_codigos["todos_os_codigos"])

	try:
		descricao = estrutura.loc[estrutura["codigo_pai"] == codigo, "descricao_pai"].values[0]
	except:
		descricao = pd.read_sql(
			text("SELECT TOP 1 TRIM(B1_DESC) descricao FROM VW_MN_SB1 B1 WHERE B1.D_E_L_E_T_ <> '*' AND B1_COD = :codigo"),
			engine, params={"codigo":codigo}
		)["descricao"].values[0]

	custos_ultima_compra = busca_custos_ultima_compra(str_codigos, data_referencia, engine)
	estrutura = traz_custos_por_produto(estrutura, custos_ultima_compra, ("ult_compra_custo_utilizado","comentario_ultima_compra"))
	custos_totais_produto = calcula_custo_total(codigo, descricao, data_referencia, estrutura, 
		["ult_compra_custo_utilizado","ult_compra_custo_utilizado_alt","custo_total_ultima_compra","comentario_ultima_compra"])
	
	custos_ultimo_fechamento = busca_custos_ultimo_fechamento(str_codigos, data_referencia, engine)
	estrutura = traz_custos_por_produto(estrutura, custos_ultimo_fechamento, ("fechamento_custo_utilizado","comentario_fechamento"))
	custos_totais_produto = calcula_custo_total(codigo, descricao, data_referencia, estrutura,
		["fechamento_custo_utilizado","fechamento_custo_utilizado_alt","custo_total_ultimo_fechamento","comentario_ultimo_fechamento"],
		custos_totais_produto)
	
	custos_medios = busca_custos_medios(str_codigos, data_referencia, engine)
	estrutura = traz_custos_por_produto(estrutura, custos_medios, ("medio_atual_custo_utilizado", "comentario_custo_medio"))
	custos_totais_produto = calcula_custo_total(codigo, descricao, data_referencia, estrutura, 
		["medio_atual_custo_utilizado","medio_atual_custo_utilizado_alt","total_pelo_custo_medio", "comentario_custo_medio"],
		custos_totais_produto)

	return estrutura, custos_totais_produto
	
	wb = Workbook()
	ws = wb.active
	ws.title = "Detalhamento"

	ws.append([
		"Cód Original","Desc Orig","Código Pai", "Desc Pai", # 1, 2, 3, 4
		"Tipo Pai", "Insumo", "Descrição Insumo", "Quant Utilizada", # 5, 6, 7, 8
		"Tipo Insumo", "Origem", "Alternativos", # 9, 10, 11
		"Últ Compra", "Últ Compra Alt", # 12, 13
		"Últ Fechamento", "Últ Fecham Alt", # 14, 15
		"Médio Atual", "Médio Atual Alt" # 16, 17
	])

	for i, row in estrutura.iterrows():
		l = i+2
		ws.cell(l, 1, codigo)
		ws.cell(l, 2, descricao)
		ws.cell(l, 3, row["codigo_pai"])
		ws.cell(l, 4, row["descricao_pai"])
		ws.cell(l, 5, row["tipo_pai"])
		ws.cell(l, 6, row["insumo"])
		ws.cell(l, 7, row["descricao_insumo"])
		ws.cell(l, 8, row["quant_utilizada"])
		ws.cell(l, 9, row["tipo_insumo"])
		ws.cell(l,10, row["origem"])
		ws.cell(l,11, row["alternativos"])

		comm = row["comentario_ultima_compra"]
		ws.cell(l,12, row["ult_compra_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_ultima_compra_alt"]
		ws.cell(l,13, row["ult_compra_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_fechamento"]
		ws.cell(l,14, row["fechamento_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_fechamento_alt"]
		ws.cell(l,15, row["fechamento_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio"]
		ws.cell(l,16, row["medio_atual_custo_utilizado"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None
		
		comm = row["comentario_custo_medio_alt"]
		ws.cell(l,17, row["medio_atual_custo_utilizado_alt"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

	tab = Table(
		displayName="tabela_estruturas",
		ref=f"A1:Q{len(estrutura) + 1}"
	)
	style = TableStyleInfo(
		name="TableStyleMedium2",
		showFirstColumn=False,
		showLastColumn=False,
		showRowStripes=True,
		showColumnStripes=False,
	)
	tab.tableStyleInfo = style
	ws.add_table(tab)

	ws2 = wb.create_sheet("Consolidado",0)
	ws2.append(["Data de Referência", "Código","Descrição","Últimas Entradas","Último Fechamento","Custo Médio"])

	formato_data = NamedStyle(name="ddmmyyyy",number_format="DD/MM/YYYY")

	for i, row in custos_totais_produto.iterrows():
		l = i+2
		ws2.cell(l, 1, data_referencia).style = formato_data
		ws2.cell(l, 2, codigo)
		ws2.cell(l, 3, descricao)

		comm = row["comentario_ultima_compra"]
		ws2.cell(l, 4, row["custo_total_ultima_compra"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_ultimo_fechamento"]
		ws2.cell(l, 5, row["custo_total_ultimo_fechamento"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

		comm = row["comentario_custo_medio"]
		ws2.cell(l, 6, row["total_pelo_custo_medio"])\
			.comment = Comment(comm, "", (comm.count("\n") + 2) * 20, max(len(lin) for lin in comm.split("\n")) * 8) if comm else None

	tab = Table(
		displayName="tabela_consolidada",
		ref=f"A1:F{len(custos_totais_produto) + 1}"
	)
	tab.tableStyleInfo = style
	ws2.add_table(tab)


	wb.save(f"estruturas_{datetime.strftime(datetime.now(),'%Y-%m-%d_%H-%M')}.xlsx")

	#startfile("relatorio.xlsx")

